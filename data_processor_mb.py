# -*- coding: utf-8 -*-
"""
data_processor.py — Module Logic QA mới.
Chứa class QALogic thực hiện các logic đối soát chuyên sâu.
"""

import pandas as pd
import numpy as np
import re
import unicodedata
import time
import json
import os
from typing import Any, List, Dict, Optional
# Global Cache to prevent redundant calculations
GLOBAL_GEO_CACHE = {}

class QALogic:
    def __init__(self):
        self.keywords = [
            "STT", "Đối tượng", "Mã vật tư", "Tuyến cáp", "Mã đối tượng", "Vị trí", "Tên đối tượng", 
            "Port", "Điểm cuối", "Điểm đầu", "Mối hàn", "Dung lượng", "Chiều dài", "Hình thức",
            "Mã kế hoạch", "Vị trí hàn nối", "Hạng mục"
        ]
        self._geocode_cache = GLOBAL_GEO_CACHE
        self._api_fail_count = 0

    def sync_design_prefixes(self, df_design: pd.DataFrame, df_bbnt_dt: pd.DataFrame, df_bbnt_han: pd.DataFrame = None) -> pd.DataFrame:
        """
        Tự động thêm prefix (ví dụ: HDG, HNI...) cho TOÀN BỘ cột A file thiết kế.
        Nếu đối tượng bắt đầu bằng P và thiếu prefix, hệ thống sẽ lấy prefix phổ biến nhất từ BBNT (Đối tượng + Hàn nối) để gán vào.
        """
        if df_design.empty:
            return df_design

        # 1. Tìm Prefix phổ biến nhất từ BBNT Đối tượng & Hàn nối
        found_prefixes = []
        pattern = re.compile(r'^([A-Z]+)P\d+', re.IGNORECASE)
        
        # --- Nguồn 1: BBNT Đối tượng ---
        col_dt = self._find_column(df_bbnt_dt, ["Đối tượng", "Mã đối tượng"])
        if col_dt:
            for val in df_bbnt_dt[col_dt].dropna().unique():
                val_str = str(val).replace(" ", "").upper()
                match = pattern.search(val_str)
                if match: found_prefixes.append(match.group(1))
        
        # --- Nguồn 2: BBNT Hàn nối ---
        if df_bbnt_han is not None and not df_bbnt_han.empty:
            col_han = self._find_column(df_bbnt_han, ["Vị trí", "Vị trí hàn", "Đối tượng"])
            if col_han:
                for val in df_bbnt_han[col_han].dropna().unique():
                    val_str = str(val).replace(" ", "").upper()
                    match = pattern.search(val_str)
                    if match: found_prefixes.append(match.group(1))

        # Xác định prefix đại diện (ưu tiên cái xuất hiện nhiều nhất)
        rep_prefix = ""
        if found_prefixes:
            from collections import Counter
            rep_prefix = Counter(found_prefixes).most_common(1)[0][0]

        # 2. Áp dụng chuẩn hóa cho các cột A, B, C, D, E, H, K theo yêu cầu User (A=0, B=1, C=2, D=3, E=4, H=7, K=10)
        target_indices = [0, 1, 2, 3, 4, 7, 10]
        
        if not df_design.empty:
            def _standardize_all(val):
                if pd.isna(val) or str(val).strip() == '':
                    return val
                
                # Làm sạch: Xóa mọi khoảng trắng, Viết hoa
                v_clean = str(val).replace(" ", "").upper()
                
                # Nếu bắt đầu bằng P (thiếu prefix) và tìm thấy prefix đại diện từ BBNT
                if v_clean.startswith('P') and rep_prefix:
                    return f"{rep_prefix}{v_clean}"
                
                return v_clean

            for idx in target_indices:
                if idx < len(df_design.columns):
                    col_name = df_design.columns[idx]
                    df_design[col_name] = df_design[col_name].apply(_standardize_all)

        return df_design

    def identify_file_type(self, df: pd.DataFrame, filename: str = "", sheet_name: str = "") -> str:
        """
        Nhận diện loại file dựa trên tên sheet trước, sau đó tên file, sau đó mới dùng cột.
        Returns: 'Form_import', 'thiet_ke', 'doi_tuong', 'TUYEN_CAP', 'han_noi', 'vat_tu', 'UNKNOWN'
        """
        fname = filename.lower()
        sname = sheet_name.lower()
        cols = [str(c).lower().strip() for c in df.columns]

        # --- Rule 0: Sheetname-based Detection (Highest Priority) ---
        if 'form_import_doi_tuong' in sname or 'form_import_doituong' in sname:
            return "Form_import"
        if 'form_import_cap' in sname:
            return "Form_import_cap"
        if 'form_import' in sname:
            return "Form_import"
        if 'form_thiet_ke' in sname or 'thiet_ke' in sname:
            return "thiet_ke"

        # --- Rule 1: Filename-based Detection (Legacy Support) ---
        if 'thiet_ke' in fname or 'bang_thiet_ke' in fname:
            return "thiet_ke"
        
        if 'formimport' in fname or 'form_import' in fname:
            return "Form_import"
            
        if 'bbnt_vattu' in fname or 'bbnt_vt' in fname:
            return "vat_tu"
            
        if 'bbnt_hannoi' in fname or 'bbnt_hn' in fname or 'hàn nối' in fname:
            return "han_noi"
            
        if 'bbnt_dt' in fname or 'bbnt_doituong' in fname:
            return "doi_tuong"
            
        if 'bbnt_tuyencap' in fname or 'bbnt_cap' in fname:
            return "TUYEN_CAP"

        # --- Rule 2: Column-based Heuristics (Fallback) ---
        has_ma_dt = any(x in cols for x in ['mã đối tượng', 'mã đt', 'ma doi tuong'])
        has_doi_tuong = any(x in cols for x in ['đối tượng', 'tên đối tượng'])
        has_port = any(x in cols for x in ['port', 'số port', 'port gpon'])
        has_tuyen = any(x in cols for x in ['tuyến cáp', 'tên tuyến'])
        has_dai = any(x in cols for x in ['chiều dài', 'chiều dài tuyến', 'length'])
        has_han = any(x in cols for x in ['vị trí hàn', 'mối hàn', 'sl mối hàn', 'vị trí'])
        has_vattu = any(x in cols for x in ['mã vật tư', 'tên vật tư'])
        has_sl_tk = any(x in cols for x in ['sl thiết kế', 'sl dự toán'])
        has_sl_tt = any(x in cols for x in ['sl thực tế', 'sl nghiệm thu'])
        has_cong_suat = any(x in cols for x in ['công suất', 'power', 'thu'])

        # 1. BBNT / TK Tuyến Cáp
        if has_tuyen and (has_dai or 'chiều dài' in str(cols)):
            return "TUYEN_CAP"
            
        # 2. BBNT / TK Hàn Nối
        if has_han and (has_sl_tt or has_sl_tk or 'mối hàn' in str(cols)):
            return "han_noi"
            
        # 3. FormImport / BBNT DT (chung đặc điểm Port + Mã)
        if (has_ma_dt or has_doi_tuong) and has_port:
            return "doi_tuong"
            
        # 4. Vật tư
        if has_vattu and (has_sl_tt or has_sl_tk):
            return "vat_tu"
            
        # 5. Thiết kế (Tổng hợp - ko có Port)
        if (has_ma_dt or has_doi_tuong) and not has_port:
             return "thiet_ke"

        return "UNKNOWN"


    def _normalize_text(self, val: Any) -> str:
        """
        Chuẩn hóa text: xóa khoảng trắng, viết hoa để đối soát công tâm.
        (Không phân biệt chữ hoa, chữ thường, khoảng trống dư thừa).
        """
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return ""
        # Xóa sạch mọi khoảng trắng và viết hoa toàn bộ
        return str(val).replace(" ", "").strip().upper()

    def _safe_num(self, val: Any) -> float:
        """
        Chuyển đổi giá trị sang float an toàn.
        Tự động bóc tách số từ chuỗi (ví dụ: '25m' -> 25.0, ' 12.5 kg ' -> 12.5).
        """
        if val is None or (isinstance(val, (float, np.float64, np.float32)) and np.isnan(val)):
            return 0.0
        
        # Nếu là số thật sự, trả về luôn
        if isinstance(val, (int, float, np.number)):
            return float(val)

        # Chuyển sang chuỗi và làm sạch đầu cuối
        s = str(val).strip()
        if not s:
            return 0.0
            
        # Thử ép kiểu trực tiếp trước (thay dấu phẩy thành chấm để float() hiểu được)
        s_dot = s.replace(',', '.')
        try:
            res = float(s_dot)
            return 0.0 if np.isnan(res) else res
        except (ValueError, TypeError):
            # Nếu thất bại, dùng Regex để tìm cụm số đầu tiên trong chuỗi
            # Hỗ trợ: số nguyên, số thập phân (dùng . hoặc ,), số âm
            # Pattern: [-+]? (dấu) \d+ (số nguyên) (?:[.,]\d+)? (phần thập phân)
            match = re.search(r'([-+]?\d+(?:[.,]\d+)?)', s)
            if match:
                num_str = match.group(1).replace(',', '.')
                try:
                    res = float(num_str)
                    return 0.0 if np.isnan(res) else res
                except:
                    pass
            return 0.0

    def _remove_diacritics(self, text: str) -> str:
        """Loại bỏ dấu tiếng Việt để so sánh fuzzy."""
        nfkd = unicodedata.normalize('NFKD', text)
        return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()

    def _find_column(self, df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        """Tìm tên cột trong DF khớp với candidates."""
        df_cols_lower = {str(c).lower().strip(): c for c in df.columns}
        
        # Exact match
        for c in candidates:
            if c.lower().strip() in df_cols_lower:
                return df_cols_lower[c.lower().strip()]
        
        # Partial match
        for c in candidates:
            for col_lower, col_orig in df_cols_lower.items():
                if c.lower() in col_lower:
                    return col_orig
        return None

    def read_excel(self, file: Any) -> Dict[str, pd.DataFrame]:
        """
        Đọc file Excel thông minh và Triệt để:
        Hệ thống "Phòng ngự 5 tầng" để hóa giải mọi lỗi file từ hệ thống xuất ra.
        Returns: Dictionary {sheet_name: DataFrame}
        """
        import io
        import pandas as pd
        import numpy as np

        def reset_f():
            if hasattr(file, 'seek'): file.seek(0)

        # 1. Lấy dữ liệu thô
        file_bytes = None
        try:
            if hasattr(file, 'read'):
                reset_f()
                file_bytes = file.read()
                reset_f()
            else: file_bytes = file
        except: return pd.DataFrame([{"LỖI_ĐỌC_FILE": "Không thể truy cập dữ liệu file."}])

        if not file_bytes: return pd.DataFrame([])

        # --- TẦNG 1: PANDAS DEFAULT (Tất cả engine) ---
        for eng in ['calamine', 'openpyxl', 'xlrd', 'pyxlsb']:
            try:
                sheets = pd.read_excel(io.BytesIO(file_bytes), header=None, sheet_name=None, engine=eng)
                if sheets: return self._process_sheets_dict(sheets)
            except: continue

        # --- TẦNG 2: HTML/XML (Dành cho file xuất hệ thống giả Excel) ---
        try:
            html_dfs = pd.read_html(io.BytesIO(file_bytes))
            if html_dfs: return self._process_sheets_dict({f"H_{i}": d for i, d in enumerate(html_dfs)})
        except: pass

        try:
            xml_df = pd.read_xml(io.BytesIO(file_bytes))
            if not xml_df.empty: return self._process_sheets_dict({"XML": xml_df})
        except: pass

        # --- TẦNG 3: MANUAL OPENPYXL (Bỏ qua Styles) ---
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True, keep_vba=False)
            sheets = {}
            for sn in wb.sheetnames:
                ws = wb[sn]
                data = [[cell.value for cell in row] for row in ws.iter_rows(values_only=True)]
                if data: sheets[sn] = pd.DataFrame(data)
            if sheets: return self._process_sheets_dict(sheets)
        except: pass

        # --- TẦNG 4: CSV/TEXT FALLBACK ---
        for enc in ['utf-8-sig', 'latin-1', 'utf-16', 'cp1252']:
            try:
                df_csv = pd.read_csv(io.BytesIO(file_bytes), sep=None, engine='python', encoding=enc, header=None)
                if not df_csv.empty and len(df_csv.columns) > 1:
                    return self._process_sheets_dict({"CSV": df_csv})
            except: continue

        # --- TẦNG 5: "XRAY" (GIẢI PHÁP CUỐI CÙNG - ĐỌC TRỰC TIẾP ZIP XML) ---
        # Dùng cho file XLSX bị lỗi Stylesheet nặng khiến Openpyxl thất bại
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                # Đọc Shared Strings (Chuỗi lặp lại)
                ss = []
                if 'xl/sharedStrings.xml' in z.namelist():
                    with z.open('xl/sharedStrings.xml') as f:
                        tree = ET.parse(f)
                        # Tìm tất cả thẻ <t> trong các thẻ <si>
                        ss = [t.text for t in tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')]
                
                # Đọc từng Sheet
                ws_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet')]
                sheets = {}
                for ws_f in ws_files:
                    with z.open(ws_f) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        rows = []
                        for r_node in root.findall('.//ns:row', ns):
                            row_vals = []
                            for c_node in r_node.findall('ns:c', ns):
                                val = ""
                                v = c_node.find('ns:v', ns)
                                if v is not None:
                                    val = v.text
                                    if c_node.get('t') == 's' and val.isdigit():
                                        idx = int(val)
                                        if idx < len(ss): val = ss[idx]
                                row_vals.append(val)
                            rows.append(row_vals)
                        if rows: sheets[ws_f] = pd.DataFrame(rows)
                if sheets: return self._process_sheets_dict(sheets)
        except Exception as e:
            final_err = str(e)

        return {"LỖI": pd.DataFrame([{"LỖI_ĐỌC_FILE": f"Hệ thống không thể giải mã file này. Lỗi sau cùng: {final_err}"}])}

    def _process_sheets_dict(self, sheets_dict: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        """Logic tìm Header và trả về dictionary tất cả các sheet hợp lệ."""
        processed_sheets = {}
        
        for sheet_name, df_full in sheets_dict.items():
            if df_full.empty: continue
            
            # Chuyển sơ bộ về string để quét Header
            df_full_str = df_full.astype(str)
            header_row = -1
            max_row_score = -1
            
            # Tìm Header trong 50 dòng đầu
            for idx, row in df_full_str.head(50).iterrows():
                row_texts = [str(v).strip().lower() for v in row.values if pd.notna(v)]
                valid_texts = [v for v in row_texts if v not in ('nan', '', 'none', 'null')]
                kw_matches = sum(1 for kw in self.keywords if any(kw.lower() in val for val in valid_texts))
                row_score = len(valid_texts) * 10 + kw_matches * 50
                if row_score > max_row_score:
                    max_row_score = row_score
                    header_row = idx
            
            # Nếu điểm số quá thấp (không giống bảng dữ liệu), bỏ qua sheet này
            # Hạ xuống 5 để chấp nhận các sheet đơn giản
            if max_row_score < 5: 
                continue
            
            # Thiết lập Header
            df = df_full.iloc[header_row:].copy()
            df.columns = df.iloc[0].values
            df = df.iloc[1:].copy()
            
            # Chuẩn hóa tên cột
            cleaned_cols = []
            seen = {}
            for i, c in enumerate(df.columns):
                c_str = str(c).strip().replace('\n', ' ').replace('  ', ' ')
                if c_str.lower() in ('nan', '', 'none', 'null'): c_str = f"Unnamed_{i}"
                if c_str in seen:
                    seen[c_str] += 1
                    c_str = f"{c_str}_{seen[c_str]}"
                else: seen[c_str] = 0
                cleaned_cols.append(c_str)
            df.columns = cleaned_cols
            
            # Làm sạch dữ liệu rỗng (xóa dòng/cột rỗng hoàn toàn)
            df = df.replace('nan', np.nan).replace('', np.nan).dropna(how='all').reset_index(drop=True)
            
            # Lọc bỏ các dòng lặp lại của chính Header (nếu có)
            if 'STT' in df.columns:
                 stt_col = df['STT']
                 if isinstance(stt_col, pd.DataFrame): stt_col = stt_col.iloc[:, 0]
                 df = df[~stt_col.astype(str).str.lower().isin(['stt', 'số tt', 'số thứ tự', 'nan', 'none'])]
            
            df = df.dropna(how='all').reset_index(drop=True)
            
            if not df.empty:
                processed_sheets[sheet_name] = df
                
        return processed_sheets

    def check_doi_tuong(self, df_import: pd.DataFrame, df_bbnt_dt: pd.DataFrame) -> pd.DataFrame:
        """
        1. Hàm check_doi_tuong(df_formimport, df_bbnt_dt):
           - Outer Join theo 'Mã đối tượng'.
           - Check: 'Mở port', 'Công suất'.
           - Cảnh báo: Thiếu đối tượng, Thừa đối tượng.
        """
    def check_doi_tuong(self, df_import: pd.DataFrame, df_bbnt_dt: pd.DataFrame, df_design: pd.DataFrame = None) -> pd.DataFrame:
        """
        1. Hàm check_doi_tuong (Refactored):
           Data Sources:
           - BBNT: df_bbnt_dt (Tab DoiTuong) -> Check Geo/Address/Status/Power.
           - Import: df_import (Tab FormImport) -> Calc Capacity.
           - Design: df_design (Tab ThietKe) -> Calc Capacity & Map Codes.
           
           Checks:
           1. BBNT: Geo (D) vs Address (E).
           2. BBNT: SP2 Power/Port Status.
           3. Cross: Design Cap (1:16->16) vs Import Cap (Sum SP1/SP2 > 0).
           4. Cross: Design Box Code (H) vs Import Object Code (O).
        """
        results = []
        
        # --- 1. PREPARE BBNT DATA (Primary View) ---
        # Col B=Đối tượng(1), G=Trạng thái(6), I=Công suất(8)
        col_obj_bbnt = self._find_column(df_bbnt_dt, ["Đối tượng", "Mã đối tượng"])
        if not col_obj_bbnt and len(df_bbnt_dt.columns) > 1: col_obj_bbnt = df_bbnt_dt.columns[1]
        
        # User: "Mở port là lấy cột G trong file đối tượng" -> Index 6
        col_stat_bbnt = self._find_column(df_bbnt_dt, ["Mở port", "Trạng thái port"])
        if not col_stat_bbnt and len(df_bbnt_dt.columns) > 6: col_stat_bbnt = df_bbnt_dt.columns[6]
        
        col_pwr_bbnt = self._find_column(df_bbnt_dt, ["Công suất", "Power"])
        if not col_pwr_bbnt and len(df_bbnt_dt.columns) > 8: col_pwr_bbnt = df_bbnt_dt.columns[8]

        if not col_obj_bbnt:
             return pd.DataFrame({"Lỗi": ["BBNT: Không tìm thấy cột Đối tượng (B)"]})

        # --- PREPARE DATA DICTS FOR CROSS CHECK ---
        # We build lookup dicts by BOTH Code and Name to ensure matching.
        # Import: Code=Col O (14), Name=Col C (2)
        import_caps = {} 
        import_name_to_code = {} # New: Map Name -> Code (Col C -> Col O)
        
        if not df_import.empty:
            has_code = len(df_import.columns) > 14
            has_name = len(df_import.columns) > 2
            
            for _, row in df_import.iterrows():
                try:
                    # Calc Cap: Sum G-L (Indices 6-11)
                    vals = row.iloc[6:12].apply(self._safe_num)
                    cap = vals[vals > 0].sum()
                    
                    # Logic SP1/SP2 Label: G-I (6-9) vs J-L (9-12)
                    sp1_sum = row.iloc[6:9].apply(self._safe_num).sum()
                    sp2_sum = row.iloc[9:12].apply(self._safe_num).sum()
                    lbl = ""
                    if sp1_sum > 0 and sp2_sum > 0: lbl = "SP1,2"
                    elif sp1_sum > 0: lbl = "SP1"
                    elif sp2_sum > 0: lbl = "SP2"
                    
                    val_data = (cap, lbl)
                    
                    c_val = None
                    n_val = None
                    
                    # Store by Code (Col O)
                    if has_code:
                        c_val = str(row.iloc[14]).strip().upper()
                        if c_val and c_val != 'NAN': 
                            import_caps[c_val] = val_data
                        
                    # Store by Name (Col C) - User fallback
                    if has_name:
                        n_val = str(row.iloc[2]).strip().upper()
                        if n_val and n_val != 'NAN': 
                            import_caps[n_val] = val_data
                            # Map Name -> Code
                            if c_val and c_val != 'NAN':
                                import_name_to_code[n_val] = c_val
                except: pass

        # Design: Code=Col H (7), Name=Col A (0)
        design_caps = {} 
        design_name_to_code = {} # New: Map Name -> Code (Col A -> Col H)
        
        if df_design is not None and not df_design.empty:
             has_box_code = len(df_design.columns) > 7
             has_box_name = len(df_design.columns) > 0
             has_sp = len(df_design.columns) > 1
             
             for _, row in df_design.iterrows():
                 try:
                     d_cap = 0
                     s_val = ""
                     if has_sp: s_val = str(row.iloc[1])
                     
                     # User Request: "2x1:8 là 16..." and "1:16 là 16..."
                     # Pass 1: Handle Multipliers (Nx1:M)
                     matches_mul = re.findall(r'(\d+)[xX]1:(\d+)', s_val)
                     for count, cap in matches_mul:
                         d_cap += int(count) * int(cap)
                         
                     # Pass 2: Handle Singles (1:M) NOT preceded by Nx
                     # Use negative lookbehind to ensure we don't count the '1:8' inside '2x1:8' again
                     # Pattern: Not preceded by digit+x/X, match 1:digits
                     matches_single = re.findall(r'(?<!\d[xX])1:(\d+)', s_val)
                     d_cap += sum(int(m) for m in matches_single)

                     if d_cap == 0:
                         d_cap = self._safe_num(s_val)

                     c_val = None
                     n_val = None
                     
                     # Store by Code (Col H)
                     if has_box_code:
                         c_val = str(row.iloc[7]).strip().upper()
                         if c_val and c_val != 'NAN': 
                             design_caps[c_val] = d_cap
                         
                     # Store by Name (Col A)
                     if has_box_name:
                         n_val = str(row.iloc[0]).strip().upper()
                         if n_val and n_val != 'NAN': 
                             design_caps[n_val] = d_cap
                             # Map Name -> Code
                             if c_val and c_val != 'NAN':
                                 design_name_to_code[n_val] = c_val
                 except: pass

        # --- PROCESS BBNT ROWS ---
        for _, row in df_bbnt_dt.iterrows():
            obj_name = row[col_obj_bbnt]
            if pd.isna(obj_name) or str(obj_name).strip() == '': continue
            
            # User Request: "Các giá trị của đối tượng chỉ lấy đến hết định dạng đối tượng (ví dụ HNIP295.0368/HO)"
            # Identify format by presence of '/' for Splitter/ODF codes. 
            # If standard materials (200...), skip? 
            # The example implies we only want rows that look like Object Codes (with /).
            if '/' not in str(obj_name):
                continue

            obj_key = str(obj_name).strip().upper()
            errors = []
            warns = []
            
            # 2. SP2 Logic (Power -22 to -10 AND Status Open)
            # Identify if object is SP2? Or apply to all? "Các đối tượng SP2 phải..."
            # Assume check applies if Name contains "SP2"? Or check all rows?
            # Let's apply to all rows where we have Power data, or if name implies SP2.
            is_sp2 = "SP2" in obj_name.upper() # Simple heuristic?
            # Actually user said "Các đối tượng SP2...". Let's assume applies if columns exist.
            
            pwr = self._safe_num(row.get(col_pwr_bbnt, -999))
            status = str(row.get(col_stat_bbnt, '')).lower()
            
            # --- Logic for Columns ---
            
            # Smart Key Match for Cross Checks (Design/Import)
            # Try exact match first, then base match (remove /HO suffix)
            base_key = obj_key.split('/')[0] if '/' in obj_key else obj_key
            
            d_cap = design_caps.get(obj_key) or design_caps.get(base_key, 0)
            
            # Import Cap is now a tuple (cap, label)
            i_data = import_caps.get(obj_key) or import_caps.get(base_key, (0, ""))
            if isinstance(i_data, (int, float)): i_data = (i_data, "") # Fallback just in case
            i_cap, i_lbl = i_data
            
            # Format as int for display if whole number, skip if NaN
            d_disp = d_cap
            if not np.isnan(d_cap):
                d_disp = int(d_cap) if d_cap == int(d_cap) else d_cap
                
            i_disp = i_cap
            if not np.isnan(i_cap):
                i_disp = int(i_cap) if i_cap == int(i_cap) else i_cap

            # 2. SP2 Check (Power/Port)
            sp2_check = "-"
            
            # Determine if we should enforce strict check (SP2) or just info (SP1 or No Cap)
            # Use i_lbl from Import if available, else name heuristic
            is_sp1_only = (i_lbl == "SP1")
            is_no_cap = (i_cap == 0)
            
            if pwr != -999: # If we have power data
                pwr_ok = (-22 <= pwr <= -10)
                port_ok = ("đã mở" in status or "open" in status or "ok" in status)
                
                if pwr_ok and port_ok:
                    sp2_check = "✅ Đạt"
                else:
                    if is_sp1_only:
                        sp2_check = f"ℹ️ (SP1) P={pwr}, St={status}"
                    elif is_no_cap:
                        # No Capacity defined in Import -> Treat as Info
                        sp2_check = f"ℹ️ (Ko DL) P={pwr}, St={status}"
                    else:
                        # SP2 and Has Capacity: Strict Error
                        sp2_check = f"❌ Lỗi (P={pwr}, St={status})"
                        if not pwr_ok: errors.append(f"Công suất {pwr} ngoài vùng [-22, -10]")
                        if not port_ok: errors.append(f"Trạng thái '{status}' chưa mở")

            # 4. Code Cross Check (Retrieved early to check for 'M' type boxes)
            design_code = design_name_to_code.get(obj_key) or design_name_to_code.get(base_key)
            import_code = import_name_to_code.get(obj_key) or import_name_to_code.get(base_key)
            
            has_m_box = False
            if (design_code and 'M' in str(design_code).upper()) or (import_code and 'M' in str(import_code).upper()):
                has_m_box = True

            # 3. Capacity Cross Check
            cap_check = "❌ 0/0"
            if d_cap > 0 and i_cap > 0:
                if d_cap == i_cap:
                    cap_check = f"✅ {d_disp}/{i_disp} {i_lbl}".strip()
                else:
                    cap_check = f"❌ {d_disp}/{i_disp} {i_lbl}".strip()
                    if not has_m_box: errors.append(f"Lệch Dlượng: TK={d_cap} vs Imp={i_cap}")
            elif d_cap > 0:
                cap_check = f"❌ {d_disp}/0"
                if not has_m_box: errors.append("Thiếu DL Import")
            elif i_cap > 0:
                cap_check = f"❌ 0/{i_disp} {i_lbl}".strip()
                if not has_m_box: errors.append("Thiếu DL TK")
            else:
                cap_check = "❌ 0/0"
                if not has_m_box: errors.append("Thiếu Dung lượng trong cả TK và Import")
            
            # If it's an M box, we don't treat capacity issues as errors
            if has_m_box:
                cap_check = cap_check.replace("❌", "✅")

            # 4. Code Cross Check (Continued)
            # Logic: Match by Object Name (BBNT Name == Design Col A == Import Col C)
            # Retrieve Code: Design Col H vs Import Col O
            
            # Get codes was done above for 'M' check.

            
            d_val = design_code if design_code else "-"
            i_val = import_code if import_code else "-"
            
            code_check = f"{d_val}/{i_val}"
            
            if design_code and import_code:
                if design_code == import_code:
                    code_check = f"✅ {code_check}"
                else:
                    code_check = f"❌ {code_check}"
                    errors.append(f"Lệch Mã hộp: TK={design_code} vs Imp={import_code}")
            elif design_code:
                code_check = f"❌ {code_check}"
                errors.append("Mã hộp có trong TK, thiếu trong Import")
            elif import_code:
                code_check = f"❌ {code_check}"
                errors.append("Mã hộp có trong Import, thiếu trong TK")
            else:
                code_check = "❌ -/-"
                errors.append("Không tìm thấy Mã hộp trong cả TK và Import")

            col_note_bbnt = self._find_column(df_bbnt_dt, ["Ghi chú", "Note"])
            
            results.append({
                "Đối tượng": obj_name,
                "Check Công suất/Mở port": sp2_check,
                "Dung lượng (Thiết kế/Import)": cap_check,
                "Mã hộp (Thiết kế/Import)": code_check,
                "Chi tiết Lỗi khác": "\n".join(["- " + str(e).strip() for e in (errors + warns)]),
                "Ghi chú": ""
            })
            
        return pd.DataFrame(results)


    def check_han_noi(self, df_tk: pd.DataFrame, df_bbnt_hannoi: pd.DataFrame, df_imp: pd.DataFrame = None) -> pd.DataFrame:
        """
        2.Hàm check_han_noi (Updated):
           - Vị trí: Cột C file Han_noi
           - SL Thực tế: Cột E file Han_noi
           - SL Thiết kế: So khớp Vị trí với Cột A file list Thiết Kế, lấy số liệu từ Cột K file Thiết Kế.
           - Cross-check: So Vị trí với Tên đối tượng trong form_import_doi_tuong.
        """
        if "LỖI_ĐỌC_FILE" in df_bbnt_hannoi.columns:
            return pd.DataFrame({"Lỗi": [f"Chi tiết (Hàn nối): {df_bbnt_hannoi['LỖI_ĐỌC_FILE'].iloc[0]}"]})
            
        if len(df_bbnt_hannoi.columns) < 5:
            avail_cols = ", ".join(list(df_bbnt_hannoi.columns))
            return pd.DataFrame({"Lỗi": [f"BBNT Hàn nối: Rất ít cột ({len(df_bbnt_hannoi.columns)} cột). Yêu cầu Vị trí (C), Thực tế (E). Các cột: {avail_cols}"]})
            
        # BBNT Hàn nối: C=2, E=4
        c_name_vitri = df_bbnt_hannoi.columns[2]
        c_name_tt = df_bbnt_hannoi.columns[4]
        
        # --- Pre-process Import Data for Cross-check ---
        imp_lookup = {}
        if df_imp is not None and not df_imp.empty:
            c_imp_obj = self._find_column(df_imp, ["Tên đối tượng", "Tên thiết bị", "Đối tượng"])
            c_imp_cause = self._find_column(df_imp, ["Nguyên nhân điều chỉnh"])
            c_imp_content = self._find_column(df_imp, ["Nội dung điều chỉnh"])
            
            if c_imp_obj and c_imp_cause and c_imp_content:
                for _, r in df_imp.iterrows():
                    o_name = str(r.get(c_imp_obj, "")).strip()
                    if o_name:
                        imp_lookup[o_name] = {
                            "cause": str(r.get(c_imp_cause, "")).strip(),
                            "content": str(r.get(c_imp_content, "")).strip()
                        }

        # Hàm chuẩn hóa Key để so khớp thiết kế
        def get_norm_key(val):
            return self._normalize_text(str(val).split('-')[0].strip())
            
        # Ánh xạ Thiết kế (Cột A -> Cột K)
        dict_tk = {}
        tk_positions_raw = {} # Để lưu tên gốc phục vụ hiển thị nếu thiếu
        if not df_tk.empty and len(df_tk.columns) > 10:
            c_tk_key = df_tk.columns[0]   # Cột A
            c_tk_val = df_tk.columns[10]  # Cột K
            for _, r in df_tk.iterrows():
                k_val = str(r.get(c_tk_key, "")).strip()
                v_tk = self._safe_num(r.get(c_tk_val, 0))
                if k_val and k_val.lower() != 'nan' and v_tk > 0:
                    norm_k = get_norm_key(k_val)
                    dict_tk[norm_k] = dict_tk.get(norm_k, 0) + v_tk
                    if norm_k not in tk_positions_raw:
                        tk_positions_raw[norm_k] = k_val
        
        # Count occurrences of Vị trí (Vị trí trùng nhau báo lỗi đỏ)
        pos_counts = df_bbnt_hannoi[c_name_vitri].astype(str).str.strip().value_counts().to_dict()
        
        results = []
        matched_tk_keys = set()
        
        # 1. Quét qua BBNT (Đề nghị)
        for _, row in df_bbnt_hannoi.iterrows():
            pos = str(row.get(c_name_vitri, "")).strip()
            if not pos or pos.lower() == 'nan':
                continue
                
            sl_tt = self._safe_num(row.get(c_name_tt, 0))
            
            # Get SL Thiết kế
            norm_pos = get_norm_key(pos)
            sl_tk = dict_tk.get(norm_pos, 0)
            if norm_pos in dict_tk:
                matched_tk_keys.add(norm_pos)
            
            diff = sl_tk - sl_tt
            status = "✅ Khớp"
            msg = []
            
            # Check Lệch số lượng hoặc thiếu
            if sl_tk == 0 and sl_tt > 0:
                status = "❌ Thiếu trong thiết kế"
                msg.append(f"Vị trí có trong Đề nghị (TT={sl_tt}) nhưng KHÔNG có trong Thiết kế")
            elif sl_tk > 0 and sl_tt == 0:
                status = "❌ Thiếu trong đề nghị"
                msg.append(f"Vị trí có trong Thiết kế (TK={sl_tk}) nhưng KHÔNG có trong Đề nghị")
            elif diff != 0:
                status = "❌ Lệch số lượng"
                msg.append(f"TK={sl_tk} - TT={sl_tt} => Lệch {diff}")
            
            # Check Trùng vị trí
            if pos_counts.get(pos, 0) > 1:
                status = "❌ Trùng vị trí"
                msg.append(f"CẢNH BÁO: Vị trí này xuất hiện {pos_counts[pos]} lần trong danh sách")
            
            # --- New Cross-check with Import file ---
            if pos in imp_lookup:
                info = imp_lookup[pos]
                if info["cause"] == "Điều chỉnh mối hàn":
                    adj_content = info['content']
                    msg.append(f"thay đổi mối hàn thành {adj_content}")
                    
                    # Nếu số thay đổi = cột SL thực tế thì cột trạng thái lỗi sẽ là lỗi thực tế (theo yêu cầu user)
                    adj_num = self._safe_num(adj_content)
                    if adj_num == sl_tt and adj_num != 0:
                        status = "❌ Lỗi thực tế"

            results.append({
                "Vị trí": pos,
                "SL Thiết kế": f"{sl_tk:g}",
                "SL đề nghị": f"{sl_tt:g}",
                "Trạng thái Lỗi": status,
                "Chi tiết": "\n".join(["- " + str(e).strip() for e in msg]),
                "Ghi chú": ""
            })
            
        # 2. Quét các vị trí có trong Thiết kế nhưng BBNT KHÔNG có
        for tk_key, tk_val in dict_tk.items():
            if tk_key not in matched_tk_keys:
                raw_name = tk_positions_raw.get(tk_key, tk_key)
                results.append({
                    "Vị trí": raw_name,
                    "SL Thiết kế": f"{tk_val:g}",
                    "SL đề nghị": "0",
                    "Trạng thái Lỗi": "❌ Thiếu trong biên bản",
                    "Chi tiết": f"- Vị trí có trong Thiết kế (TK={tk_val}) nhưng KHÔNG có trong BBNT Đề nghị",
                    "Ghi chú": ""
                })
                
        return pd.DataFrame(results)

    def check_tuyen_cap(self, df_tk: pd.DataFrame, df_tuyencap: pd.DataFrame, df_imp_cap: pd.DataFrame = None) -> pd.DataFrame:
        """
        3. Hàm check_tuyen_cap (Updated):
           - BBNT (tuyen_cap): Target = Col D (Điểm cuối).
           - Design (tk): Source = Col A (Tên đối tượng).
           - Match Logic: BBNT['Điểm cuối'] (normalized) == TK['Tên đối tượng'].
           - Comparisons:
             1. Dung lượng: Col F (BBNT) vs Col C (TK).
             2. Loại cáp: Col G (BBNT) [Treo->CT, Ngam->CC] vs Col D (TK).
             3. Chiều dài: Col I (BBNT) vs Col E (TK).
        """
        # Check for catastrophic read error
        if "LỖI_ĐỌC_FILE" in df_tuyencap.columns:
            return pd.DataFrame({"Lỗi": [f"Chi tiết (Tuyến cáp): {df_tuyencap['LỖI_ĐỌC_FILE'].iloc[0]}"]})
            
        # --- 1. Identify Columns (Hardcode theo yêu cầu user B=1, C=2, D=3, F=5, G=6, I=8) ---
        col_tuyen_bbnt = df_tuyencap.columns[1] if len(df_tuyencap.columns) > 1 else self._find_column(df_tuyencap, ["Tuyến cáp"])
        col_start_bbnt = df_tuyencap.columns[2] if len(df_tuyencap.columns) > 2 else self._find_column(df_tuyencap, ["Điểm đầu", "Tên điểm đầu", "From"])
        col_end_bbnt   = df_tuyencap.columns[3] if len(df_tuyencap.columns) > 3 else self._find_column(df_tuyencap, ["Điểm cuối"])
        col_cap_bbnt   = df_tuyencap.columns[5] if len(df_tuyencap.columns) > 5 else self._find_column(df_tuyencap, ["Dung lượng cáp", "Dung lượng"])
        col_type_bbnt  = df_tuyencap.columns[6] if len(df_tuyencap.columns) > 6 else self._find_column(df_tuyencap, ["Hình thức sử dụng", "Hình thức"])
        col_len_h_bbnt = df_tuyencap.columns[7] if len(df_tuyencap.columns) > 7 else self._find_column(df_tuyencap, ["Chiều dài tuyến cáp"])
        col_len_bbnt   = df_tuyencap.columns[8] if len(df_tuyencap.columns) > 8 else self._find_column(df_tuyencap, ["Chiều dài dự toán"])
        col_cs_dau     = df_tuyencap.columns[9] if len(df_tuyencap.columns) > 9 else self._find_column(df_tuyencap, ["Chỉ số đầu"])
        col_cs_cuoi    = df_tuyencap.columns[10] if len(df_tuyencap.columns) > 10 else self._find_column(df_tuyencap, ["Chỉ số cuối"])

        if not col_tuyen_bbnt or not col_end_bbnt:
            avail_cols = ", ".join(list(df_tuyencap.columns[:5]))
            return pd.DataFrame({"Lỗi": [f"BBNT Tuyến cáp: Rất ít cột ({len(df_tuyencap.columns)} cột). Các cột hiện tại: {avail_cols}"]})

        if df_tk is None or df_tk.empty:
            df_tk = pd.DataFrame(columns=["Tên đối tượng", "Cột rỗng 1", "DL cáp", "Loại cáp", "Số lượng"])

        # TK - User specified: A=TenDoiTuong(0), C=DLCap(2), D=LoaiCap(3), E=SoLuong(4)
        col_obj_tk  = self._find_column(df_tk, ["Tên đối tượng", "Đối tượng"])
        if not col_obj_tk and len(df_tk.columns) > 0: col_obj_tk = df_tk.columns[0]
        
        col_cap_tk  = self._find_column(df_tk, ["DL cáp", "Dung lượng"])
        if not col_cap_tk and len(df_tk.columns) > 2: col_cap_tk = df_tk.columns[2]
        
        col_type_tk = self._find_column(df_tk, ["Loại cáp", "Loại"])
        if not col_type_tk and len(df_tk.columns) > 3: col_type_tk = df_tk.columns[3]
        
        col_len_tk  = self._find_column(df_tk, ["Số lượng", "Chiều dài"])
        if not col_len_tk and len(df_tk.columns) > 4: col_len_tk = df_tk.columns[4]
        
        if not col_obj_tk:
             return pd.DataFrame({"Lỗi": ["Design: Không tìm thấy cột Tên đối tượng (A)"]})

        # --- PREPARE IMPORT CABLE DATA FOR CROSS CHECK ---
        imp_cap_lookup = {}
        if df_imp_cap is not None and not df_imp_cap.empty:
            c_imp_tuyen = self._find_column(df_imp_cap, ["Tên cáp", "Tuyến cáp"])
            c_imp_cause = self._find_column(df_imp_cap, ["Nguyên nhân điều chỉnh"])
            c_imp_cause_detail = self._find_column(df_imp_cap, ["Nguyên nhân chi tiết"])
            c_imp_content = self._find_column(df_imp_cap, ["Nội dung điều chỉnh"])
            
            if c_imp_tuyen and c_imp_cause:
                for _, r in df_imp_cap.iterrows():
                    t_name_raw = str(r.get(c_imp_tuyen, "")).strip()
                    t_name_norm = self._normalize_text(t_name_raw)
                    cause_val = str(r.get(c_imp_cause, "")).strip()
                    
                    if t_name_norm and cause_val and cause_val.lower() != 'nan' and cause_val != '':
                        detail = str(r.get(c_imp_cause_detail, "")).strip()
                        if detail.lower() == 'nan': detail = ""
                        content = str(r.get(c_imp_content, "")).strip()
                        if content.lower() == 'nan': content = ""

                        # Build display string starting with 'Nguyên nhân điều chỉnh' value
                        info_parts = []
                        if detail: info_parts.append(detail)
                        if content: info_parts.append(content)
                        
                        combined_info = f"{cause_val}: " + " - ".join(info_parts)
                        imp_cap_lookup[t_name_norm] = combined_info

        # --- 2. Prepare Match Keys ---
        # User Request 1: "Dữ liệu theo hàng chỉ lấy đến hết định dạng tuyến cáp (ví dụ như HNIP295.0369/CO"
        # -> Filter: Only keep rows where 'Tuyến cáp' (Col B) contains '/'
        df_bbnt = df_tuyencap.copy()
        
        # Filter rows based on format (contains '/')
        if col_tuyen_bbnt:
            df_bbnt = df_bbnt[df_bbnt[col_tuyen_bbnt].astype(str).str.contains('/', na=False)]
        
        df_tk = df_tk.copy()
        
        # User Request: Lấy phần trước dấu gạch ngang (Code)
        def extract_endpoint(val):
            s = str(val).split('-')[0]
            return self._normalize_text(s.strip())

        df_bbnt['_key'] = df_bbnt[col_end_bbnt].apply(extract_endpoint)
        df_tk['_key'] = df_tk[col_obj_tk].apply(extract_endpoint)

        # Merge - Left Join on BBNT to keep list of Cables
        # "Dữ liệu hiện ra tab tuyến cáp được lấy từ cột tuyến cáp (cột B) trong file tuyen_cap"
        merged = pd.merge(df_bbnt, df_tk, on='_key', how='left', suffixes=('_BBNT', '_TK'))
        
        # Helper for resolving merged column names
        def _g(r, col, suffix):
            # Check original column name
            if col in r: return r[col]
            # Check suffixed column name
            if str(col)+suffix in r: return r[str(col)+suffix]
            return None

        results = []
        for _, row in merged.iterrows():
            tuyen_name = _g(row, col_tuyen_bbnt, '_BBNT')
            start_val_raw = _g(row, col_start_bbnt, '_BBNT')
            start_val = extract_endpoint(start_val_raw) if pd.notna(start_val_raw) and str(start_val_raw).strip() != '' else ""
            
            # Skip empty rows if any
            if pd.isna(tuyen_name) or str(tuyen_name).strip() == '':
                continue
                
            errors = []
            
            # Check Match
            obj_tk_val = _g(row, col_obj_tk, '_TK')
            if pd.isna(obj_tk_val):  # If merge failed
                 pass 

            # 1. Capacity (F vs C)
            cap_bbnt = self._safe_num(_g(row, col_cap_bbnt, '_BBNT'))
            cap_tk = self._safe_num(_g(row, col_cap_tk, '_TK'))
            if cap_bbnt != cap_tk:
                errors.append(f"Lệch Dlượng: TT={cap_bbnt} vs TK={cap_tk}")

            # 2. Type (G vs D) -> Mapping tương đương Treo=CT, Ngầm=CC
            # _normalize_text đã viết hoa toàn bộ và xóa khoảng trắng
            type_bbnt_raw = self._normalize_text(row.get(col_type_bbnt))
            type_tk_raw = self._normalize_text(row.get(col_type_tk))
            
            def get_standard_type(t_raw):
                if not t_raw: return ""
                if any(x in t_raw for x in ["TREO", "CT", "CAPTREO"]): return "CT"
                if any(x in t_raw for x in ["NGAM", "NGẦM", "CC", "CAPNGAM", "CAPNGẦM"]): return "CC"
                return t_raw

            std_bbnt = get_standard_type(type_bbnt_raw)
            std_tk = get_standard_type(type_tk_raw)
            
            if std_bbnt != std_tk:
                 if not std_tk and std_bbnt:
                     errors.append(f"Thiếu Loại cáp TK (TT={type_bbnt_raw})")
                 elif std_tk:
                     errors.append(f"Lệch Loại: TT={type_bbnt_raw} vs TK={type_tk_raw}")

            # 3. Length (Col H, Col I in BBNT vs Col E in TK)
            len_h_bbnt = self._safe_num(row.get(col_len_h_bbnt))
            len_i_bbnt = self._safe_num(row.get(col_len_bbnt))
            len_tk = self._safe_num(row.get(col_len_tk))
            
            # --- So sánh nội bộ BBNT (H vs I) ---
            is_len_tc_err = False
            if len_h_bbnt > len_i_bbnt:
                is_len_tc_err = True
                errors.append(f"C.dài thi công ({len_h_bbnt}) > Dự toán tool ({len_i_bbnt})")
            elif len_i_bbnt > 0 and len_h_bbnt < len_i_bbnt * 0.5:
                is_len_tc_err = True
                errors.append(f"C.dài thi công ({len_h_bbnt}) < 50% Dự toán tool ({len_i_bbnt})")

            # --- So sánh BBNT vs Thiết kế (I vs TK) ---
            if abs(len_i_bbnt - len_tk) > 1:
                errors.append(f"Lệch Chiều dài file: BBNT_TK={len_i_bbnt} vs TK_Gốc={len_tk}")

            status = "✅ Khớp"
            if errors: 
                status = "❌ Sai lệch"
            
            # If no Design found at all for this key?
            if pd.isna(row.get(col_cap_tk)) and pd.isna(row.get(col_len_tk)):
                 status = "⚠️ Không tìm thấy TK"
                 errors = [f"Không tìm thấy đối tượng '{row['_key']}' trong file TK"]

            # --- Cell-level markers for styling ---
            cap_marker = "✅" if cap_bbnt == cap_tk else "❌"
            type_marker = "✅" if std_bbnt == std_tk else "❌"
            
            # Length logic: match external (I vs TK)
            is_len_err = (abs(len_i_bbnt - len_tk) > 1)
            len_marker = "❌" if is_len_err else "✅"
            len_tc_marker = "❌" if is_len_tc_err else "✅"

            # --- Cross-check with Form_import_cap ---
            tuyen_name_norm = self._normalize_text(tuyen_name)
            if tuyen_name_norm in imp_cap_lookup:
                errors.append(f"{imp_cap_lookup[tuyen_name_norm]}")

            col_note_tc = self._find_column(df_bbnt, ["Ghi chú", "Note"])
            
            cs_dau_raw = row.get(col_cs_dau) if col_cs_dau else None
            cs_cuoi_raw = row.get(col_cs_cuoi) if col_cs_cuoi else None
            cs_dau_val = f"{self._safe_num(cs_dau_raw):g}" if pd.notna(cs_dau_raw) and str(cs_dau_raw).strip() != '' else ""
            cs_cuoi_val = f"{self._safe_num(cs_cuoi_raw):g}" if pd.notna(cs_cuoi_raw) and str(cs_cuoi_raw).strip() != '' else ""

            results.append({
                "Tuyến cáp": tuyen_name,
                "Điểm đầu": start_val,
                "Điểm cuối (Key)": row['_key'],
                "Dung lượng (TT/TK)": f"{cap_marker} {int(cap_bbnt)} / {int(cap_tk)}",
                "Loại (TT/TK)": f"{type_marker} {type_bbnt_raw} / {type_tk_raw}",
                "C.dài thi công": f"{len_tc_marker} {len_h_bbnt:g}",
                "Dự toán tool/ Thiết kế": f"{len_marker} {len_i_bbnt:g} / {len_tk:g}",
                "Chỉ số đầu": cs_dau_val,
                "Chỉ số cuối": cs_cuoi_val,
                "Trạng thái Lỗi": status,
                "Chi tiết": "\n".join(["- " + str(e).strip() for e in errors]),
                "Ghi chú": ""
            })
            
        return pd.DataFrame(results)

    def check_vat_tu(self, df_bbnt_vattu: pd.DataFrame, df_bbnt_dt: pd.DataFrame, df_bbnt_tuyencap: pd.DataFrame) -> pd.DataFrame:
        """
        4. Hàm check_vat_tu (Updated):
        - Bỏ gom nhóm, xử lý từng dòng trong BBNT Vật tư.
        - SL đối chiếu = Tổng SL thực tế (Cột J) file doi_tuong + Tổng SL thực tế (Cột E) file tuyen_cap theo Mã Vật Tư.
        - SL Thực tế = SL thực tế (Cột H) file vat_tu.
        - Hiển thị Tình trạng hàng, bỏ Đơn vị tính.
        """
        # --- 1. Xác định các cột trong df_bbnt_vattu (File Vật Tư) ---
        col_code_vt = self._find_column(df_bbnt_vattu, ["Mã vật tư", "Mã VT"])
        if not col_code_vt and len(df_bbnt_vattu.columns) > 1: col_code_vt = df_bbnt_vattu.columns[1]
        
        col_name_vt = self._find_column(df_bbnt_vattu, ["Tên vật tư", "Tên VT", "Nội dung", "Diễn giải"])
        if not col_name_vt and len(df_bbnt_vattu.columns) > 2: col_name_vt = df_bbnt_vattu.columns[2]
        
        # SL thực tế trong vat_tu (User: Cột H -> index 7)
        col_sl_nt = self._find_column(df_bbnt_vattu, ["SL Thực tế", "SL Nghiệm thu", "Khối lượng NT", "Thực tế"])
        if not col_sl_nt and len(df_bbnt_vattu.columns) > 7: col_sl_nt = df_bbnt_vattu.columns[7]
        
        # Tình trạng hàng trong vat_tu
        col_tinhtrang = self._find_column(df_bbnt_vattu, ["Tình trạng hàng", "Tình trạng", "Chất lượng"])
        if not col_tinhtrang and len(df_bbnt_vattu.columns) > 8: col_tinhtrang = df_bbnt_vattu.columns[8] # Fallback bừa nếu ko có

        if not col_code_vt:
             return pd.DataFrame({"Lỗi": ["Không tìm thấy cột 'Mã vật tư' trong file BBNT Vật tư"]})

        # --- 2. Tính tổng SL Thiết Kế từ doi_tuong và tuyen_cap ---
        design_caps = {} # Sum mapping: Code -> Total SL Thiết kế

        # Từ file doi_tuong: Mã vật tư ở cột B (index 1), SL thực tế bù đắp ở cột J (index 9)
        if df_bbnt_dt is not None and not df_bbnt_dt.empty:
            current_code = ""
            for i in range(len(df_bbnt_dt)):
                try:
                    code_raw = str(df_bbnt_dt.iloc[i, 1]).strip().upper()
                    # Bỏ qua dòng tiêu đề, rác
                    if code_raw == 'MÃ VẬT TƯ' or code_raw == 'MÃ VT':
                        continue

                    if code_raw and code_raw != 'NAN' and code_raw != 'NONE':
                        current_code = code_raw

                    if current_code:
                        val = self._safe_num(df_bbnt_dt.iloc[i, 9])
                        design_caps[current_code] = design_caps.get(current_code, 0) + val
                except IndexError:
                    pass # Bỏ qua nếu dòng không đủ cột

        # Từ file tuyen_cap: Mã vật tư ở cột B (index 1), SL thực tế ở cột E (index 4)
        if df_bbnt_tuyencap is not None and not df_bbnt_tuyencap.empty:
            current_code = ""
            for i in range(len(df_bbnt_tuyencap)):
                try:
                    code_raw = str(df_bbnt_tuyencap.iloc[i, 1]).strip().upper()
                    if code_raw == 'MÃ VẬT TƯ' or code_raw == 'MÃ VT':
                        continue

                    if code_raw and code_raw != 'NAN' and code_raw != 'NONE':
                        current_code = code_raw

                    if current_code:
                        val = self._safe_num(df_bbnt_tuyencap.iloc[i, 4])
                        design_caps[current_code] = design_caps.get(current_code, 0) + val
                except IndexError:
                    pass

        # --- 3. Duyệt từng dòng trong BBNT Vật tư và Gom nhóm theo (Mã, Kho) ---
        col_kho = self._find_column(df_bbnt_vattu, ["Kho", "Tên kho", "Kho hàng"])
        # Nếu ko tìm thấy theo tên, thử lấy cột D (index 3) - giả định theo form mẫu
        if not col_kho and len(df_bbnt_vattu.columns) > 3: col_kho = df_bbnt_vattu.columns[3]

        grouped_vt = {} # (Code, Kho) -> {name, sl_nt, tinh_trang_list}
        for _, row in df_bbnt_vattu.iterrows():
            code_raw = row.get(col_code_vt)
            if pd.isna(code_raw) or str(code_raw).strip() == '': continue
            
            code = str(code_raw).strip().upper()
            if code == 'NAN': continue
            
            kho_val = str(row.get(col_kho, '')).strip() if col_kho else ""
            if pd.isna(row.get(col_kho)): kho_val = ""

            name = str(row.get(col_name_vt, ''))
            tinh_trang = str(row.get(col_tinhtrang, '')) if col_tinhtrang else ''
            if pd.isna(row.get(col_tinhtrang)): tinh_trang = ''
            
            s_nt = self._safe_num(row.get(col_sl_nt, 0))
            
            group_key = (code, kho_val)
            if group_key not in grouped_vt:
                grouped_vt[group_key] = {
                    "code": code,
                    "kho": kho_val,
                    "name": name,
                    "sl_nt": s_nt,
                    "tinh_trang": [tinh_trang] if tinh_trang else []
                }
            else:
                grouped_vt[group_key]["sl_nt"] += s_nt
                if tinh_trang and tinh_trang not in grouped_vt[group_key]["tinh_trang"]:
                    grouped_vt[group_key]["tinh_trang"].append(tinh_trang)

        results = []
        for key, data in grouped_vt.items():
            code = data["code"]
            kho = data["kho"]
            name = data["name"]
            s_nt = data["sl_nt"]
            tinh_trang_str = ", ".join(data["tinh_trang"])
            s_tk = design_caps.get(code, 0.0)
            
            # Compare
            status = "✅ Khớp"
            details = []
            
            diff = s_tk - s_nt
            if abs(diff) > 0.001:
                status = "❌ Lệch SL"
                details.append(f"TK={s_tk:g} - NT={s_nt:g} => Lệch {diff:g}")

            results.append({
                "Kho": kho,
                "Mã vật tư": code,
                "Tên vật tư": name,
                "Tình trạng hàng": tinh_trang_str,
                "SL đối chiếu": f"{round(s_tk, 1)}",
                "SL Nghiệm thu": f"{round(s_nt, 1)}",
                "Trạng thái Lỗi": status,
                "Chi tiết": "\n".join(["- " + str(e).strip() for e in details]),
                "Ghi chú": ""
            })
             
        if not results:
            return pd.DataFrame({"Lỗi": ["Không tìm thấy dữ liệu hợp lệ trong file BBNT Vật tư"]})
            
        # Sắp xếp theo Kho
        results.sort(key=lambda x: (str(x["Kho"]), str(x["Mã vật tư"])))

        return pd.DataFrame(results)
             
        if not results:
            return pd.DataFrame({"Lỗi": ["Không tìm thấy dữ liệu hợp lệ trong file BBNT Vật tư"]})
            
        return pd.DataFrame(results)

    def check_design_capacity(self, df_imp: pd.DataFrame, df_tk: pd.DataFrame) -> pd.DataFrame:
        """
        New Check: So sánh Form_import và thiet_ke (Dung lượng).
        - Imp: Tổng SP1.x, SP2.x (số port có dữ liệu).
        - Tk: Parse 'Loại bộ chia' (1:16 -> 16, 1:8 -> 8).
        """
        # --- 1. Calculate Import Usage ---
        # Find columns SP1.* and SP2.*
        sp1_cols = [c for c in df_imp.columns if 'sp1.' in str(c).lower()]
        sp2_cols = [c for c in df_imp.columns if 'sp2.' in str(c).lower()]
        
        # Calculate used ports
        def count_used(row, cols):
            cnt = 0
            for c in cols:
                val = str(row[c]).strip().lower()
                if val and val != 'nan' and val != '0' and val != 'none':
                    cnt += 1
            return cnt

        df_imp = df_imp.copy()
        df_imp['Used_SP1'] = df_imp.apply(lambda r: count_used(r, sp1_cols), axis=1)
        df_imp['Used_SP2'] = df_imp.apply(lambda r: count_used(r, sp2_cols), axis=1)

        key_imp = self._find_column(df_imp, ["Mã đối tượng", "Mã ĐT"])
        if not key_imp: return pd.DataFrame({"Lỗi": ["Không tìm thấy Mã ĐT trong FormImport"]})
        
        df_imp['_key'] = df_imp[key_imp].apply(self._normalize_text)
        
        # --- 2. Get Design Capacity ---
        key_tk = self._find_column(df_tk, ["Mã đối tượng", "Mã ĐT"])
        col_type = self._find_column(df_tk, ["Loại bộ chia", "Loại thiết bị", "Loại"])
        
        if not key_tk: return pd.DataFrame({"Lỗi": ["Không tìm thấy Mã ĐT trong Thiết kế"]})
        
        df_tk = df_tk.copy()
        df_tk['_key'] = df_tk[key_tk].apply(self._normalize_text)
        
        # Parse Capacity from Type
        def parse_cap(val):
            s = str(val).lower()
            if '1:16' in s: return 16
            if '1:8' in s: return 8
            if '1:4' in s: return 4
            if '1:2' in s: return 2
            if '1:32' in s: return 32
            if '1:64' in s: return 64
            return 0 # Unknown
            
        cap_val = 0
        if col_type:
            df_tk['_DesignCap'] = df_tk[col_type].apply(parse_cap)
        else:
            df_tk['_DesignCap'] = 0

        # --- 3. Merge & Compare ---
        merged = pd.merge(df_imp, df_tk, on='_key', how='inner', suffixes=('_IMP', '_TK'))
        
        results = []
        for _, row in merged.iterrows():
            ma_dt = row['_key']
            used_1 = row['Used_SP1']
            used_2 = row['Used_SP2']
            design_cap = row.get('_DesignCap', 0)
            
            # Logic user requested:
            # "thiet_ke thì dung lượng theo nguyên tắc 1:16; là bộ chia cấp 1 16 port, 1:16 là bộ chia cấp 2 16 port"
            # So Design Cap applies to BOTH Level 1 and Level 2? 
            # Or depends on device type?
            # Assuming the 'Loại bộ chia' column defines the max capacity for THAT device instance.
            # If used > design => Error.
            
            status = "✅ Khớp"
            msgs = []
            
            # Check SP1 (Assuming this device handles SP1?)
            # Or is SP1/SP2 distinct devices?
            # Usually FormImport has one row per 'Splitter' representing one Box.
            # If Box has splits 1:16, max is 16.
            # Total used = SP1_used + SP2_used?
            # Or SP1 columns are for Level 1 splitters and SP2 for Level 2?
            # If the row is ONE Object, it usually is ONE Splitter level.
            # But if columns SP1... and SP2... exist in SAME row, maybe it lists connections?
            # Let's check max(Used_SP1, Used_SP2) vs Design?
            # Or Sum?
            # User said: "dung lượng bộ chia cấp 1 là tổng... SP1.x, cấp 2 là tổng ... SP2.x".
            # And Design "1:16 là bộ chia cấp 1... 1:16 là bộ chia cấp 2".
            # This implies the Design Type tells us if it is Lv1 or Lv2.
            # If Design says "1:16", and used (sum) > 16 => Error.
            # I will compare MAX(Used_SP1, Used_SP2) against DesignCap for safety, assumption is one device per row.
            
            total_used = used_1 + used_2
            # Refine: if SP1 and SP2 both present, maybe it's 2 devices?
            # But merge is 1:1.
            # I'll compare Total Used ports vs Design Cap.
            
            if design_cap > 0:
                if total_used > design_cap:
                    status = "❌ Quá tải (Used > Design)"
                    msgs.append(f"Used={total_used} > Design={design_cap}")
                elif total_used == 0:
                    status = "⚠️ Không có thuê bao (Empty)"
            else:
                 msgs.append("Không xác định được loại bộ chia trong TK")
            
            if status != "✅ Khớp":
                 results.append({
                    "Mã ĐT": ma_dt,
                    "Used SP1": f"{used_1}",
                    "Used SP2": f"{used_2}",
                    "Design Cap": f"{design_cap}",
                    "Trạng thái Lỗi": status,
                    "Chi tiết": "\n".join(["- " + str(e).strip() for e in msgs])
                })
        
        if not results:
             return pd.DataFrame({"Kết quả": ["OK - Không thấy lỗi dung lượng"]})
             
        return pd.DataFrame(results)

    def calculate_total_design_length(self, df_tk):
        """Calculates total design count (rows with integer format) from column 'Số lượng' or index 4."""
        if df_tk is None or df_tk.empty:
            return 0
            
        # Try finding 'Số lượng' (Col E)
        col = self._find_column(df_tk, ["Số lượng", "Chiều dài"])
        if not col and len(df_tk.columns) > 4:
            col = df_tk.columns[4]
            
        if not col:
            return 0
            
        count = 0
        for val in df_tk[col]:
            if pd.isna(val): continue
            s_val = str(val).strip()
            
            # Check if it has an integer format (ignoring letters/spaces)
            # Remove all non-digit characters to see if anything remains
            digits = "".join(filter(str.isdigit, s_val))
            if digits:
                count += 1
        return count

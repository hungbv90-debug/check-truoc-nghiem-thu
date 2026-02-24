# -*- coding: utf-8 -*-
"""
data_processor.py — Module Logic QA mới.
Chứa class QALogic thực hiện các logic đối soát chuyên sâu.
"""

import pandas as pd
import numpy as np
import re
import unicodedata
import time
import json
import os
from typing import Optional, List, Any, Dict, Tuple
from urllib.request import urlopen, Request
from urllib.error import URLError, HTTPError

try:
    import geopandas as gpd
    from shapely.geometry import Point
    HAS_GEOPANDAS = True
except ImportError:
    HAS_GEOPANDAS = False

# Global Cache to prevent redundant API calls across UI re-runs
GLOBAL_GEO_CACHE = {}
# Global Offline Boundaries (GeoDataFrame)
GLOBAL_OFFLINE_GDF = None
# Disable geocoding globally if API fails to prevent infinite spinning
GLOBAL_GEO_DISABLED = False

class QALogic:
    def __init__(self):
        self.keywords = [
            "STT", "Đối tượng", "Mã vật tư", "Tuyến cáp", "Mã đối tượng", "Vị trí", "Tên đối tượng", 
            "Port", "Điểm cuối", "Điểm đầu", "Mối hàn", "Dung lượng", "Chiều dài", "Hình thức",
            "Mã kế hoạch", "Vị trí hàn nối"
        ]
        self._geocode_cache = GLOBAL_GEO_CACHE
        self._api_fail_count = 0
        self._offline_gdf = self._load_offline_boundaries()

    def identify_file_type(self, df: pd.DataFrame, filename: str = "") -> str:
        """
        Nhận diện loại file dựa trên tên file trước, sau đó mới dùng cột.
        Returns: 'Form_import', 'thiet_ke', 'doi_tuong', 'TUYEN_CAP', 'han_noi', 'vat_tu', 'UNKNOWN'
        """
        fname = filename.lower()
        cols = [str(c).lower().strip() for c in df.columns]

        # --- Rule 1: Filename-based Detection (Priority) ---
        if 'thiet_ke' in fname or 'bang_thiet_ke' in fname:
            return "thiet_ke"
        
        if 'formimport' in fname or 'form_import' in fname:
            return "Form_import"
            
        if 'bbnt_vattu' in fname or 'bbnt_vt' in fname:
            return "vat_tu"
            
        if 'bbnt_hannoi' in fname or 'bbnt_hn' in fname or 'hàn nối' in fname:
            return "han_noi"
            
        if 'bbnt_dt' in fname or 'bbnt_doituong' in fname:
            return "doi_tuong"
            
        if 'bbnt_tuyencap' in fname or 'bbnt_cap' in fname:
            return "TUYEN_CAP"

        # --- Rule 2: Column-based Heuristics (Fallback) ---
        has_ma_dt = any(x in cols for x in ['mã đối tượng', 'mã đt', 'ma doi tuong'])
        has_doi_tuong = any(x in cols for x in ['đối tượng', 'tên đối tượng'])
        has_port = any(x in cols for x in ['port', 'số port', 'port gpon'])
        has_tuyen = any(x in cols for x in ['tuyến cáp', 'tên tuyến'])
        has_dai = any(x in cols for x in ['chiều dài', 'chiều dài tuyến', 'length'])
        has_han = any(x in cols for x in ['vị trí hàn', 'mối hàn', 'sl mối hàn', 'vị trí'])
        has_vattu = any(x in cols for x in ['mã vật tư', 'tên vật tư'])
        has_sl_tk = any(x in cols for x in ['sl thiết kế', 'sl dự toán'])
        has_sl_tt = any(x in cols for x in ['sl thực tế', 'sl nghiệm thu'])
        has_cong_suat = any(x in cols for x in ['công suất', 'power', 'thu'])

        # 1. BBNT / TK Tuyến Cáp
        if has_tuyen and (has_dai or 'chiều dài' in str(cols)):
            return "TUYEN_CAP"
            
        # 2. BBNT / TK Hàn Nối
        if has_han and (has_sl_tt or has_sl_tk or 'mối hàn' in str(cols)):
            return "han_noi"
            
        # 3. FormImport / BBNT DT (chung đặc điểm Port + Mã)
        if (has_ma_dt or has_doi_tuong) and has_port:
            return "doi_tuong"
            
        # 4. Vật tư
        if has_vattu and (has_sl_tt or has_sl_tk):
            return "vat_tu"
            
        # 5. Thiết kế (Tổng hợp - ko có Port)
        if (has_ma_dt or has_doi_tuong) and not has_port:
             return "thiet_ke"

        return "UNKNOWN"


    def _normalize_text(self, val: Any) -> str:
        """Chuẩn hóa text: strip, lower."""
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return ""
        return str(val).strip().lower()

    def _safe_num(self, val: Any) -> float:
        """Chuyển đổi giá trị sang float an toàn."""
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return 0.0
        try:
            s = str(val).strip().replace(',', '.')
            # Xử lý trường hợp "1.234.567" -> xoá hết chấm trừ cái cuối? 
            # Quy ước đơn giản: nếu có nhiều chấm thì chỉ giữ lại chấm thập phân cuối hoặc coi như bỏ qua
            # Ở đây dùng logic đơn giản: replace , -> . rồi ép kiểu
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    def _remove_diacritics(self, text: str) -> str:
        """Loại bỏ dấu tiếng Việt để so sánh fuzzy."""
        nfkd = unicodedata.normalize('NFKD', text)
        return ''.join(c for c in nfkd if not unicodedata.combining(c)).lower().strip()

    def _bigdatacloud_reverse(self, lat: float, lon: float) -> dict:
        """
        Gọi BigDataCloud free client API (không cần API key, không rate limit).
        Returns: dict chứa {province, district, ward} hoặc None nếu lỗi.
        Tốc độ: ~50-100ms/request, nhanh gấp 5-10x geopy/Nominatim.
        """
        cache_key = f"{lat:.6f},{lon:.6f}"
        
        if cache_key in self._geocode_cache:
            return self._geocode_cache[cache_key]
        
        try:
            url = f"https://api.bigdatacloud.net/data/reverse-geocode-client?latitude={lat}&longitude={lon}&localityLanguage=vi"
            req = Request(url, headers={"User-Agent": "FTEL-QA-Check/2.0"})
            
            with urlopen(req, timeout=5) as resp:
                data = json.loads(resp.read().decode('utf-8'))
            
            result = {"province": "", "district": "", "ward": ""}
            result["province"] = data.get("principalSubdivision", "") or data.get("city", "")
            result["district"] = data.get("locality", "")
            
            admin_list = data.get("localityInfo", {}).get("administrative", [])
            for item in admin_list:
                order = item.get("order", 0)
                name = item.get("name", "")
                if not name:
                    continue
                if order == 4 and not result["province"]:
                    result["province"] = name
                elif order == 6:
                    result["district"] = name
                elif order == 8:
                    result["ward"] = name
            
            self._geocode_cache[cache_key] = result
            return result
                
        except (URLError, HTTPError, TimeoutError, Exception) as e:
            self._geocode_cache[cache_key] = None
            return None

    def _load_offline_boundaries(self):
        """Tải dữ liệu ranh giới ngoại tuyến nếu có."""
        global GLOBAL_OFFLINE_GDF
        if not HAS_GEOPANDAS: return None
        if GLOBAL_OFFLINE_GDF is not None: return GLOBAL_OFFLINE_GDF
        
        # Đường dẫn tìm kiếm file ranh giới (ưu tiên độ chi tiết giảm dần)
        possible_paths = [
            "gis_data/vietnam_boundaries.geojson",
            "gis_data/vietnam_districts.geojson",
            "gis_data/vietnam_provinces.geojson"
        ]
        all_gdfs = []
        for p in possible_paths:
            if os.path.exists(p):
                try:
                    gdf = gpd.read_file(p)
                    # Chuẩn hóa về WGS84 (EPSG:4326) để khớp với tọa độ GPS thông dụng
                    if gdf.crs is None:
                        gdf.set_crs(epsg=4326, inplace=True)
                    elif gdf.crs != "EPSG:4326":
                        gdf = gdf.to_crs(epsg=4326)
                    all_gdfs.append(gdf)
                except: pass
        
        if all_gdfs:
            # Gộp các file ranh giới lại
            GLOBAL_OFFLINE_GDF = pd.concat(all_gdfs, ignore_index=True)
            return GLOBAL_OFFLINE_GDF
        return None

    def _offline_geocode_check(self, lat: float, lon: float) -> Optional[dict]:
        """Sử dụng GeoPandas để tìm thông tin hành chính offline."""
        if not HAS_GEOPANDAS or self._offline_gdf is None:
            return None
        
        try:
            p = Point(lon, lat)
            match = self._offline_gdf[self._offline_gdf.contains(p)]
            if not match.empty:
                row = match.iloc[0]
                res = {"province": "", "district": "", "ward": ""}
                
                # Heuristic mapping thông minh hơn dựa trên các bộ dữ liệu phổ biến
                for col in match.columns:
                    c_low = col.lower()
                    val = str(row[col])
                    if val.lower() in ['nan', 'none', '']: continue
                    
                    # 1. Check Tỉnh/TP
                    if any(x in c_low for x in ['name_1', 'province', 'tinh', 'city', 'tp']):
                        res["province"] = val
                    # 2. Check Quận/Huyện
                    elif any(x in c_low for x in ['name_2', 'district', 'huyen', 'quan']):
                        res["district"] = val
                    # 3. Check Phường/Xã
                    elif any(x in c_low for x in ['name_3', 'ward', 'xa', 'phuong']):
                        res["ward"] = val
                    # 4. Fallback cho Highcharts hoặc các dataset đơn giản
                    elif c_low == 'name' and not res["province"]:
                        res["province"] = val

                if res["province"] or res["district"]:
                    return res
        except: pass
        return None

    def _reverse_geocode_check(self, lat: float, lon: float, address_str: str) -> Tuple[str, str]:
        """
        Kiểm tra tọa độ vs địa chỉ. 
        Ưu tiên: Offline GIS (GeoPandas) -> Online API (BigDataCloud).
        """
        if not (8 <= lat <= 24 and 102 <= lon <= 110):
            return f"❌ Ngoài vùng VN ({lat:.4f},{lon:.4f})", f"Tọa độ nằm ngoài lãnh thổ Việt Nam"
        
        global GLOBAL_GEO_DISABLED
        
        geo_result = None
        used_api = False
        source = "Offline"
        
        # 1. Thử Offline GIS trước (Tốc độ cực nhanh, ko cần mạng)
        geo_result = self._offline_geocode_check(lat, lon)
        
        # 2. Nếu ko có dữ liệu offline, thử Online API
        if not geo_result and not GLOBAL_GEO_DISABLED:
            geo_result = self._bigdatacloud_reverse(lat, lon)
            if geo_result:
                used_api = True
                source = "API"
            else:
                self._api_fail_count += 1
                if self._api_fail_count >= 3:
                    GLOBAL_GEO_DISABLED = True
        
        if geo_result:
            geo_province = geo_result.get("province", "")
            geo_district = geo_result.get("district", "")
            geo_ward = geo_result.get("ward", "")
            
            addr_norm = self._remove_diacritics(address_str)
            
            issues = []
            matched_parts = []
            
            if geo_province:
                prov_norm = self._remove_diacritics(geo_province)
                prov_short = re.sub(r'^(thanh pho|tinh|tp\.|t\.)\s*', '', prov_norm).strip()
                prov_parts = prov_short.split()
                if prov_short and (prov_short in addr_norm or prov_norm in addr_norm or (len(prov_parts) >= 2 and all(p in addr_norm for p in prov_parts))):
                    matched_parts.append(f"T/TP: {geo_province} ✓")
                else:
                    issues.append(f"Tỉnh/TP: {source}='{geo_province}' không khớp ĐC")
            
            if geo_district:
                dist_norm = self._remove_diacritics(geo_district)
                dist_short = re.sub(r'^(quan|huyen|thi xa|tx\.|q\.)\s*', '', dist_norm).strip()
                dist_parts = dist_short.split()
                if dist_short and (dist_short in addr_norm or dist_norm in addr_norm or (len(dist_parts) >= 2 and all(p in addr_norm for p in dist_parts))):
                    matched_parts.append(f"Q/H: {geo_district} ✓")
                else:
                    issues.append(f"Quận/Huyện: {source}='{geo_district}' không khớp ĐC")
            
            if geo_ward:
                ward_norm = self._remove_diacritics(geo_ward)
                ward_short = re.sub(r'^(phuong|xa|thi tran|p\.)\s*', '', ward_norm).strip()
                ward_parts = ward_short.split()
                if ward_short and (ward_short in addr_norm or ward_norm in addr_norm or (len(ward_parts) >= 2 and all(p in addr_norm for p in ward_parts))):
                    matched_parts.append(f"P/X: {geo_ward} ✓")
                else:
                    issues.append(f"Phường/Xã: {source}='{geo_ward}' không khớp ĐC")
            
            geocode_summary = f"[{geo_ward}, {geo_district}, {geo_province}]"
            
            if not issues:
                return f"✅ Khớp vị trí {geocode_summary}", "; ".join(matched_parts)
            elif len(issues) <= 1 and matched_parts:
                detail = "; ".join(issues + matched_parts)
                return f"⚠️ Lệch nhẹ {geocode_summary}", detail
            else:
                detail = "; ".join(issues)
            return f"❌ Sai vị trí {geocode_summary}", detail
        
        return "❌ Khớp lỗi", "Không thể xác định vị trí"

    def _find_column(self, df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        """Tìm tên cột trong DF khớp với candidates."""
        df_cols_lower = {str(c).lower().strip(): c for c in df.columns}
        
        # Exact match
        for c in candidates:
            if c.lower().strip() in df_cols_lower:
                return df_cols_lower[c.lower().strip()]
        
        # Partial match
        for c in candidates:
            for col_lower, col_orig in df_cols_lower.items():
                if c.lower() in col_lower:
                    return col_orig
        return None

    def read_excel(self, file: Any) -> pd.DataFrame:
        """
        Đọc file Excel thông minh và Triệt để:
        Hệ thống "Phòng ngự 5 tầng" để hóa giải mọi lỗi file từ hệ thống xuất ra (Protected View, Corrupt Stylesheet...).
        """
        import io
        import pandas as pd
        import numpy as np

        def reset_f():
            if hasattr(file, 'seek'): file.seek(0)

        # 1. Lấy dữ liệu thô
        file_bytes = None
        try:
            if hasattr(file, 'read'):
                reset_f()
                file_bytes = file.read()
                reset_f()
            else: file_bytes = file
        except: return pd.DataFrame([{"LỖI_ĐỌC_FILE": "Không thể truy cập dữ liệu file."}])

        if not file_bytes: return pd.DataFrame([])

        # --- TẦNG 1: PANDAS DEFAULT (Tất cả engine) ---
        for eng in ['calamine', 'openpyxl', 'xlrd', 'pyxlsb']:
            try:
                sheets = pd.read_excel(io.BytesIO(file_bytes), header=None, sheet_name=None, engine=eng)
                if sheets: return self._process_sheets_dict(sheets)
            except: continue

        # --- TẦNG 2: HTML/XML (Dành cho file xuất hệ thống giả Excel) ---
        try:
            html_dfs = pd.read_html(io.BytesIO(file_bytes))
            if html_dfs: return self._process_sheets_dict({f"H_{i}": d for i, d in enumerate(html_dfs)})
        except: pass

        try:
            xml_df = pd.read_xml(io.BytesIO(file_bytes))
            if not xml_df.empty: return self._process_sheets_dict({"XML": xml_df})
        except: pass

        # --- TẦNG 3: MANUAL OPENPYXL (Bỏ qua Styles) ---
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True, keep_vba=False)
            sheets = {}
            for sn in wb.sheetnames:
                ws = wb[sn]
                data = [[cell.value for cell in row] for row in ws.iter_rows(values_only=True)]
                if data: sheets[sn] = pd.DataFrame(data)
            if sheets: return self._process_sheets_dict(sheets)
        except: pass

        # --- TẦNG 4: CSV/TEXT FALLBACK ---
        for enc in ['utf-8-sig', 'latin-1', 'utf-16', 'cp1252']:
            try:
                df_csv = pd.read_csv(io.BytesIO(file_bytes), sep=None, engine='python', encoding=enc, header=None)
                if not df_csv.empty and len(df_csv.columns) > 1:
                    return self._process_sheets_dict({"CSV": df_csv})
            except: continue

        # --- TẦNG 5: "XRAY" (GIẢI PHÁP CUỐI CÙNG - ĐỌC TRỰC TIẾP ZIP XML) ---
        # Dùng cho file XLSX bị lỗi Stylesheet nặng khiến Openpyxl thất bại
        try:
            import zipfile
            import xml.etree.ElementTree as ET
            with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
                # Đọc Shared Strings (Chuỗi lặp lại)
                ss = []
                if 'xl/sharedStrings.xml' in z.namelist():
                    with z.open('xl/sharedStrings.xml') as f:
                        tree = ET.parse(f)
                        # Tìm tất cả thẻ <t> trong các thẻ <si>
                        ss = [t.text for t in tree.findall('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')]
                
                # Đọc từng Sheet
                ws_files = [f for f in z.namelist() if f.startswith('xl/worksheets/sheet')]
                sheets = {}
                for ws_f in ws_files:
                    with z.open(ws_f) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        ns = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
                        rows = []
                        for r_node in root.findall('.//ns:row', ns):
                            row_vals = []
                            for c_node in r_node.findall('ns:c', ns):
                                val = ""
                                v = c_node.find('ns:v', ns)
                                if v is not None:
                                    val = v.text
                                    if c_node.get('t') == 's' and val.isdigit():
                                        idx = int(val)
                                        if idx < len(ss): val = ss[idx]
                                row_vals.append(val)
                            rows.append(row_vals)
                        if rows: sheets[ws_f] = pd.DataFrame(rows)
                if sheets: return self._process_sheets_dict(sheets)
        except Exception as e:
            final_err = str(e)

        return pd.DataFrame([{"LỖI_ĐỌC_FILE": f"Hệ thống không thể giải mã file này. Lỗi sau cùng: {final_err}"}])

    def _process_sheets_dict(self, sheets_dict: Dict[str, pd.DataFrame]) -> pd.DataFrame:
        """Logic tìm Header và chọn Sheet tốt nhất (Dùng chung cho mọi tầng đọc)."""
        best_df = pd.DataFrame()
        best_sheet_score = -1
        
        for sheet_name, df_full in sheets_dict.items():
            if df_full.empty: continue
            df_full = df_full.astype(str)
            header_row = -1
            max_row_score = -1
            
            # Tìm Header trong 50 dòng đầu
            for idx, row in df_full.head(50).iterrows():
                row_texts = [str(v).strip().lower() for v in row.values if pd.notna(v)]
                valid_texts = [v for v in row_texts if v not in ('nan', '', 'none', 'null')]
                kw_matches = sum(1 for kw in self.keywords if any(kw.lower() in val for val in valid_texts))
                row_score = len(valid_texts) * 10 + kw_matches * 50
                if row_score > max_row_score:
                    max_row_score = row_score
                    header_row = idx
            
            if header_row == -1: continue
            
            df = df_full.iloc[header_row:].copy()
            df.columns = df.iloc[0].values
            df = df.iloc[1:].copy()
            
            # Chuẩn hóa cột
            cleaned_cols = []
            seen = {}
            for i, c in enumerate(df.columns):
                c_str = str(c).strip().replace('\n', ' ').replace('  ', ' ')
                if c_str.lower() in ('nan', '', 'none', 'null'): c_str = f"Unnamed_{i}"
                if c_str in seen:
                    seen[c_str] += 1
                    c_str = f"{c_str}_{seen[c_str]}"
                else: seen[c_str] = 0
                cleaned_cols.append(c_str)
            df.columns = cleaned_cols
            
            # Làm sạch dữ liệu rỗng
            df = df.replace('nan', np.nan).replace('', np.nan).dropna(how='all').reset_index(drop=True)
            if 'STT' in df.columns:
                 stt_col = df['STT']
                 if isinstance(stt_col, pd.DataFrame): stt_col = stt_col.iloc[:, 0]
                 df = df[~stt_col.astype(str).str.lower().isin(['stt', 'số tt', 'số thứ tự', 'nan', 'none'])]
            df = df.dropna(how='all').reset_index(drop=True)
            
            score = max_row_score * 1000 + len(df)
            if score > best_sheet_score and not df.empty:
                best_sheet_score = score
                best_df = df
        return best_df

    def check_doi_tuong(self, df_import: pd.DataFrame, df_bbnt_dt: pd.DataFrame) -> pd.DataFrame:
        """
        1. Hàm check_doi_tuong(df_formimport, df_bbnt_dt):
           - Outer Join theo 'Mã đối tượng'.
           - Check: 'Mở port', 'Công suất'.
           - Cảnh báo: Thiếu đối tượng, Thừa đối tượng.
        """
    def check_doi_tuong(self, df_import: pd.DataFrame, df_bbnt_dt: pd.DataFrame, df_design: pd.DataFrame = None) -> pd.DataFrame:
        """
        1. Hàm check_doi_tuong (Refactored):
           Data Sources:
           - BBNT: df_bbnt_dt (Tab DoiTuong) -> Check Geo/Address/Status/Power.
           - Import: df_import (Tab FormImport) -> Calc Capacity.
           - Design: df_design (Tab ThietKe) -> Calc Capacity & Map Codes.
           
           Checks:
           1. BBNT: Geo (D) vs Address (E).
           2. BBNT: SP2 Power/Port Status.
           3. Cross: Design Cap (1:16->16) vs Import Cap (Sum SP1/SP2 > 0).
           4. Cross: Design Box Code (H) vs Import Object Code (O).
        """
        results = []
        
        # --- 1. PREPARE BBNT DATA (Primary View) ---
        # Col B=Đối tượng(1), D=Tọa độ(3), E=Địa chỉ(4), G=Trạng thái(6), I=Công suất(8)
        col_obj_bbnt = self._find_column(df_bbnt_dt, ["Đối tượng", "Mã đối tượng"])
        if not col_obj_bbnt and len(df_bbnt_dt.columns) > 1: col_obj_bbnt = df_bbnt_dt.columns[1]
        
        col_geo_bbnt = self._find_column(df_bbnt_dt, ["Tọa độ", "GPS"])
        if not col_geo_bbnt and len(df_bbnt_dt.columns) > 3: col_geo_bbnt = df_bbnt_dt.columns[3]
        
        col_addr_bbnt = self._find_column(df_bbnt_dt, ["Địa chỉ", "Khu vực"])
        if not col_addr_bbnt and len(df_bbnt_dt.columns) > 4: col_addr_bbnt = df_bbnt_dt.columns[4]
        
        col_bg_bbnt = self._find_column(df_bbnt_dt, ["Bàn giao", "Bg"])
        
        # User: "Mở port là lấy cột G trong file đối tượng" -> Index 6
        col_stat_bbnt = self._find_column(df_bbnt_dt, ["Mở port", "Trạng thái port"])
        if not col_stat_bbnt and len(df_bbnt_dt.columns) > 6: col_stat_bbnt = df_bbnt_dt.columns[6]
        
        col_pwr_bbnt = self._find_column(df_bbnt_dt, ["Công suất", "Power"])
        if not col_pwr_bbnt and len(df_bbnt_dt.columns) > 8: col_pwr_bbnt = df_bbnt_dt.columns[8]

        if not col_obj_bbnt:
             return pd.DataFrame({"Lỗi": ["BBNT: Không tìm thấy cột Đối tượng (B)"]})

        # --- PREPARE DATA DICTS FOR CROSS CHECK ---
        # We build lookup dicts by BOTH Code and Name to ensure matching.
        # Import: Code=Col O (14), Name=Col C (2)
        import_caps = {} 
        import_name_to_code = {} # New: Map Name -> Code (Col C -> Col O)
        
        if not df_import.empty:
            has_code = len(df_import.columns) > 14
            has_name = len(df_import.columns) > 2
            
            for _, row in df_import.iterrows():
                try:
                    # Calc Cap: Sum G-L (Indices 6-11)
                    vals = pd.to_numeric(row.iloc[6:12], errors='coerce').fillna(0)
                    cap = vals[vals > 0].sum()
                    
                    # Logic SP1/SP2 Label: G-I (6-9) vs J-L (9-12)
                    sp1_sum = pd.to_numeric(row.iloc[6:9], errors='coerce').fillna(0).sum()
                    sp2_sum = pd.to_numeric(row.iloc[9:12], errors='coerce').fillna(0).sum()
                    lbl = ""
                    if sp1_sum > 0 and sp2_sum > 0: lbl = "SP1,2"
                    elif sp1_sum > 0: lbl = "SP1"
                    elif sp2_sum > 0: lbl = "SP2"
                    
                    val_data = (cap, lbl)
                    
                    c_val = None
                    n_val = None
                    
                    # Store by Code (Col O)
                    if has_code:
                        c_val = str(row.iloc[14]).strip().upper()
                        if c_val and c_val != 'NAN': 
                            import_caps[c_val] = val_data
                        
                    # Store by Name (Col C) - User fallback
                    if has_name:
                        n_val = str(row.iloc[2]).strip().upper()
                        if n_val and n_val != 'NAN': 
                            import_caps[n_val] = val_data
                            # Map Name -> Code
                            if c_val and c_val != 'NAN':
                                import_name_to_code[n_val] = c_val
                except: pass

        # Design: Code=Col H (7), Name=Col A (0)
        design_caps = {} 
        design_name_to_code = {} # New: Map Name -> Code (Col A -> Col H)
        
        if df_design is not None and not df_design.empty:
             has_box_code = len(df_design.columns) > 7
             has_box_name = len(df_design.columns) > 0
             has_sp = len(df_design.columns) > 1
             
             for _, row in df_design.iterrows():
                 try:
                     d_cap = 0
                     s_val = ""
                     if has_sp: s_val = str(row.iloc[1])
                     
                     # User Request: "2x1:8 là 16..." and "1:16 là 16..."
                     # Pass 1: Handle Multipliers (Nx1:M)
                     matches_mul = re.findall(r'(\d+)[xX]1:(\d+)', s_val)
                     for count, cap in matches_mul:
                         d_cap += int(count) * int(cap)
                         
                     # Pass 2: Handle Singles (1:M) NOT preceded by Nx
                     # Use negative lookbehind to ensure we don't count the '1:8' inside '2x1:8' again
                     # Pattern: Not preceded by digit+x/X, match 1:digits
                     matches_single = re.findall(r'(?<!\d[xX])1:(\d+)', s_val)
                     d_cap += sum(int(m) for m in matches_single)

                     if d_cap == 0 and s_val.isdigit():
                         try:
                             d_cap = int(s_val)
                         except: pass

                     c_val = None
                     n_val = None
                     
                     # Store by Code (Col H)
                     if has_box_code:
                         c_val = str(row.iloc[7]).strip().upper()
                         if c_val and c_val != 'NAN': 
                             design_caps[c_val] = d_cap
                         
                     # Store by Name (Col A)
                     if has_box_name:
                         n_val = str(row.iloc[0]).strip().upper()
                         if n_val and n_val != 'NAN': 
                             design_caps[n_val] = d_cap
                             # Map Name -> Code
                             if c_val and c_val != 'NAN':
                                 design_name_to_code[n_val] = c_val
                 except: pass

        # --- PROCESS BBNT ROWS ---
        for _, row in df_bbnt_dt.iterrows():
            obj_name = row[col_obj_bbnt]
            if pd.isna(obj_name) or str(obj_name).strip() == '': continue
            
            # User Request: "Các giá trị của đối tượng chỉ lấy đến hết định dạng đối tượng (ví dụ HNIP295.0368/HO)"
            # Identify format by presence of '/' for Splitter/ODF codes. 
            # If standard materials (200...), skip? 
            # The example implies we only want rows that look like Object Codes (with /).
            if '/' not in str(obj_name):
                continue

            obj_key = str(obj_name).strip().upper()
            errors = []
            warns = []
            
            # 1. Geo (D) vs Address (E) check
            geo = str(row.get(col_geo_bbnt, '')).strip()
            addr = str(row.get(col_addr_bbnt, '')).strip()
            # "trong file đối tượng hãy so sánh tọa độ (cột D) nắm trong địa chỉ phường/huyện..."
            # This implies checking if Lat/Lon maps to Address? Complex.
            # Simplified: Check if Address contains meaningful info?
            # Or if user meant "nằm trong" literally (string containment)? Unlikely for coords.
            # Let's check Basic Consistency: valid Geo format.
            if geo and not addr: warns.append("Có Tọa độ thiếu Địa chỉ")
            elif addr and not geo: warns.append("Có Địa chỉ thiếu Tọa độ")
            
            # 2. SP2 Logic (Power -22 to -10 AND Status Open)
            # Identify if object is SP2? Or apply to all? "Các đối tượng SP2 phải..."
            # Assume check applies if Name contains "SP2"? Or check all rows?
            # Let's apply to all rows where we have Power data, or if name implies SP2.
            is_sp2 = "SP2" in obj_name.upper() # Simple heuristic?
            # Actually user said "Các đối tượng SP2...". Let's assume applies if columns exist.
            
            pwr = self._safe_num(row.get(col_pwr_bbnt, -999))
            status = str(row.get(col_stat_bbnt, '')).lower()
            
            # --- Logic for Columns ---
            
            # Smart Key Match for Cross Checks (Design/Import)
            # Try exact match first, then base match (remove /HO suffix)
            base_key = obj_key.split('/')[0] if '/' in obj_key else obj_key
            
            d_cap = design_caps.get(obj_key) or design_caps.get(base_key, 0)
            
            # Import Cap is now a tuple (cap, label)
            i_data = import_caps.get(obj_key) or import_caps.get(base_key, (0, ""))
            if isinstance(i_data, (int, float)): i_data = (i_data, "") # Fallback just in case
            i_cap, i_lbl = i_data
            
            # Format as int for display if whole number
            d_disp = int(d_cap) if d_cap == int(d_cap) else d_cap
            i_disp = int(i_cap) if i_cap == int(i_cap) else i_cap

            # 1. Location Check: Reverse Geocode → So khớp Phường/Xã + Tỉnh/TP
            loc_check = "❌ Thiếu TT"
            if geo and addr:
                try:
                    # Extract numbers. Assuming format like "21.09, 105.51"
                    parts = geo.replace(',', ' ').split()
                    nums = [float(p) for p in parts if p.replace('.','',1).isdigit()]
                    if len(nums) >= 2:
                        lat, lon = nums[0], nums[1]
                        # Swap if lat > lon (common mistake, in VN lat is ~21, lon ~105)
                        if lat > 80 and lon < 40: lat, lon = lon, lat
                        
                        # Step 1: Kiểm tra vùng VN cơ bản
                        if not (8 <= lat <= 24 and 102 <= lon <= 110):
                            loc_check = f"❌ Ngoài vùng VN ({lat},{lon})"
                            warns.append(f"Tọa độ nghi ngờ: {lat},{lon}")
                        else:
                            # Step 2: Reverse Geocode → so khớp phường/xã, tỉnh/TP với địa chỉ
                            loc_check, loc_detail = self._reverse_geocode_check(lat, lon, addr)
                            if loc_detail:
                                if '❌' in loc_check or 'Sai' in loc_check:
                                    errors.append(loc_detail)
                                elif '⚠️' in loc_check:
                                    warns.append(loc_detail)
                    else:
                        loc_check = "❌ Sai định dạng"
                except:
                    loc_check = "❌ Lỗi parse"
            
            # 2. SP2 Check (Power/Port)
            sp2_check = "-"
            
            # Determine if we should enforce strict check (SP2) or just info (SP1 or No Cap)
            # Use i_lbl from Import if available, else name heuristic
            is_sp1_only = (i_lbl == "SP1")
            is_no_cap = (i_cap == 0)
            
            if pwr != -999: # If we have power data
                pwr_ok = (-22 <= pwr <= -10)
                port_ok = ("đã mở" in status or "open" in status or "ok" in status)
                
                if pwr_ok and port_ok:
                    sp2_check = "✅ Đạt"
                else:
                    if is_sp1_only:
                        sp2_check = f"ℹ️ (SP1) P={pwr}, St={status}"
                    elif is_no_cap:
                        # No Capacity defined in Import -> Treat as Info
                        sp2_check = f"ℹ️ (Ko DL) P={pwr}, St={status}"
                    else:
                        # SP2 and Has Capacity: Strict Error
                        sp2_check = f"❌ Lỗi (P={pwr}, St={status})"
                        if not pwr_ok: errors.append(f"Công suất {pwr} ngoài vùng [-22, -10]")
                        if not port_ok: errors.append(f"Trạng thái '{status}' chưa mở")

            # 3. Capacity Cross Check
            cap_check = "-"
            if d_cap > 0 and i_cap > 0:
                if d_cap == i_cap:
                    cap_check = f"✅ {d_disp}/{i_disp} {i_lbl}".strip()
                else:
                    cap_check = f"❌ {d_disp}/{i_disp} {i_lbl}".strip()
                    errors.append(f"Lệch Dlượng: TK={d_cap} vs Imp={i_cap}")
            elif d_cap > 0:
                cap_check = f"⚠️ {d_disp}/0"
                # warns.append("Thiếu DL Import") # Reduce noise
            elif i_cap > 0:
                cap_check = f"⚠️ 0/{i_disp} {i_lbl}".strip()
                # warns.append("Thiếu DL TK")

            # 4. Code Cross Check
            # Logic: Match by Object Name (BBNT Name == Design Col A == Import Col C)
            # Retrieve Code: Design Col H vs Import Col O
            
            # Get Design Code using BBNT object name (or base key)
            design_code = design_name_to_code.get(obj_key) or design_name_to_code.get(base_key)
            # Get Import Code using BBNT object name (or base key)
            import_code = import_name_to_code.get(obj_key) or import_name_to_code.get(base_key)
            
            d_val = design_code if design_code else "-"
            i_val = import_code if import_code else "-"
            
            code_check = f"{d_val}/{i_val}"
            
            if design_code and import_code:
                if design_code == import_code:
                    code_check = f"✅ {code_check}"
                else:
                    code_check = f"❌ {code_check}"
                    errors.append(f"Lệch Mã hộp: TK={design_code} vs Imp={import_code}")
            elif design_code:
                code_check = f"⚠️ {code_check}"
                warns.append("Mã hộp có trong TK, thiếu trong Import")
            elif import_code:
                code_check = f"⚠️ {code_check}"
                warns.append("Mã hộp có trong Import, thiếu trong TK")
            else:
                code_check = "❌ -/-"
                errors.append("Không tìm thấy Mã hộp trong cả TK và Import")

            results.append({
                "Đối tượng": obj_name,
                "Kiểm tra Vị trí": loc_check,
                "Check Công suất/Mở port": sp2_check,
                "Dung lượng (Thiết kế/Import)": cap_check,
                "Mã hộp (Thiết kế/Import)": code_check,
                "Chi tiết": "; ".join(errors + warns)
            })
            
        return pd.DataFrame(results)


    def check_han_noi(self, df_tk: pd.DataFrame, df_bbnt_hannoi: pd.DataFrame) -> pd.DataFrame:
        """
        2.Hàm check_han_noi (Updated):
           - Vị trí: Cột C file Han_noi
           - SL Thực tế: Cột E file Han_noi
           - SL Thiết kế: So khớp Vị trí với Cột A file list Thiết Kế, lấy số liệu từ Cột K file Thiết Kế.
        """
        if "LỖI_ĐỌC_FILE" in df_bbnt_hannoi.columns:
            return pd.DataFrame({"Lỗi": [f"Chi tiết (Hàn nối): {df_bbnt_hannoi['LỖI_ĐỌC_FILE'].iloc[0]}"]})
            
        if len(df_bbnt_hannoi.columns) < 5:
            avail_cols = ", ".join(list(df_bbnt_hannoi.columns))
            return pd.DataFrame({"Lỗi": [f"BBNT Hàn nối: Rất ít cột ({len(df_bbnt_hannoi.columns)} cột). Yêu cầu Vị trí (C), Thực tế (E). Các cột: {avail_cols}"]})
            
        # BBNT Hàn nối: C=2, E=4
        c_name_vitri = df_bbnt_hannoi.columns[2]
        c_name_tt = df_bbnt_hannoi.columns[4]
        
        # Hàm chuẩn hóa Key để so khớp
        def get_norm_key(val):
            return self._normalize_text(str(val).split('-')[0].strip())
            
        # Ánh xạ Thiết kế (Cột A -> Cột K)
        dict_tk = {}
        if not df_tk.empty and len(df_tk.columns) > 10:
            c_tk_key = df_tk.columns[0]   # Cột A
            c_tk_val = df_tk.columns[10]  # Cột K
            for _, r in df_tk.iterrows():
                k_val = str(r.get(c_tk_key, "")).strip()
                v_tk = self._safe_num(r.get(c_tk_val, 0))
                if k_val and k_val.lower() != 'nan':
                    norm_k = get_norm_key(k_val)
                    dict_tk[norm_k] = dict_tk.get(norm_k, 0) + v_tk
        
        results = []
        for _, row in df_bbnt_hannoi.iterrows():
            pos = str(row.get(c_name_vitri, "")).strip()
            if not pos or pos.lower() == 'nan':
                continue
                
            sl_tt = self._safe_num(row.get(c_name_tt, 0))
            
            # Get SL Thiết kế
            norm_pos = get_norm_key(pos)
            sl_tk = dict_tk.get(norm_pos, 0)
            
            diff = sl_tk - sl_tt
            status = "✅ Khớp"
            msg = []
            
            if diff != 0:
                status = "❌ Lệch số lượng"
                msg.append(f"TK={sl_tk} - TT={sl_tt} => Lệch {diff}")
            
            results.append({
                "Vị trí": pos,
                "SL Thiết kế": sl_tk,
                "SL Thực tế": sl_tt,
                "Trạng thái Lỗi": status,
                "Chi tiết": "; ".join(msg)
            })
            
        return pd.DataFrame(results)

    def check_tuyen_cap(self, df_tk: pd.DataFrame, df_tuyencap: pd.DataFrame) -> pd.DataFrame:
        """
        3. Hàm check_tuyen_cap (Updated):
           - BBNT (tuyen_cap): Target = Col D (Điểm cuối).
           - Design (tk): Source = Col A (Tên đối tượng).
           - Match Logic: BBNT['Điểm cuối'] (normalized) == TK['Tên đối tượng'].
           - Comparisons:
             1. Dung lượng: Col F (BBNT) vs Col C (TK).
             2. Loại cáp: Col G (BBNT) [Treo->CT, Ngam->CC] vs Col D (TK).
             3. Chiều dài: Col I (BBNT) vs Col E (TK).
        """
        # Check for catastrophic read error
        if "LỖI_ĐỌC_FILE" in df_tuyencap.columns:
            return pd.DataFrame({"Lỗi": [f"Chi tiết (Tuyến cáp): {df_tuyencap['LỖI_ĐỌC_FILE'].iloc[0]}"]})
            
        # --- 1. Identify Columns (Hardcode theo yêu cầu user B=1, C=2, D=3, F=5, G=6, I=8) ---
        col_tuyen_bbnt = df_tuyencap.columns[1] if len(df_tuyencap.columns) > 1 else self._find_column(df_tuyencap, ["Tuyến cáp"])
        col_start_bbnt = df_tuyencap.columns[2] if len(df_tuyencap.columns) > 2 else self._find_column(df_tuyencap, ["Điểm đầu", "Tên điểm đầu", "From"])
        col_end_bbnt   = df_tuyencap.columns[3] if len(df_tuyencap.columns) > 3 else self._find_column(df_tuyencap, ["Điểm cuối"])
        col_cap_bbnt   = df_tuyencap.columns[5] if len(df_tuyencap.columns) > 5 else self._find_column(df_tuyencap, ["Dung lượng cáp", "Dung lượng"])
        col_type_bbnt  = df_tuyencap.columns[6] if len(df_tuyencap.columns) > 6 else self._find_column(df_tuyencap, ["Hình thức sử dụng", "Hình thức"])
        col_len_bbnt   = df_tuyencap.columns[8] if len(df_tuyencap.columns) > 8 else self._find_column(df_tuyencap, ["Chiều dài dự toán"])

        if not col_tuyen_bbnt or not col_end_bbnt:
            avail_cols = ", ".join(list(df_tuyencap.columns[:5]))
            return pd.DataFrame({"Lỗi": [f"BBNT Tuyến cáp: Rất ít cột ({len(df_tuyencap.columns)} cột). Các cột hiện tại: {avail_cols}"]})

        # TK - User specified: A=TenDoiTuong(0), C=DLCap(2), D=LoaiCap(3), E=SoLuong(4)
        col_obj_tk  = self._find_column(df_tk, ["Tên đối tượng", "Đối tượng"])
        if not col_obj_tk and len(df_tk.columns) > 0: col_obj_tk = df_tk.columns[0]
        
        col_cap_tk  = self._find_column(df_tk, ["DL cáp", "Dung lượng"])
        if not col_cap_tk and len(df_tk.columns) > 2: col_cap_tk = df_tk.columns[2]
        
        col_type_tk = self._find_column(df_tk, ["Loại cáp", "Loại"])
        if not col_type_tk and len(df_tk.columns) > 3: col_type_tk = df_tk.columns[3]
        
        col_len_tk  = self._find_column(df_tk, ["Số lượng", "Chiều dài"])
        if not col_len_tk and len(df_tk.columns) > 4: col_len_tk = df_tk.columns[4]
        
        if not col_obj_tk:
             return pd.DataFrame({"Lỗi": ["Design: Không tìm thấy cột Tên đối tượng (A)"]})

        # --- 2. Prepare Match Keys ---
        # User Request 1: "Dữ liệu theo hàng chỉ lấy đến hết định dạng tuyến cáp (ví dụ như HNIP295.0369/CO"
        # -> Filter: Only keep rows where 'Tuyến cáp' (Col B) contains '/'
        df_bbnt = df_tuyencap.copy()
        
        # Filter rows based on format (contains '/')
        if col_tuyen_bbnt:
            df_bbnt = df_bbnt[df_bbnt[col_tuyen_bbnt].astype(str).str.contains('/', na=False)]
        
        df_tk = df_tk.copy()
        
        # User Request: Lấy phần trước dấu gạch ngang (Code)
        def extract_endpoint(val):
            s = str(val).split('-')[0]
            return self._normalize_text(s.strip())

        df_bbnt['_key'] = df_bbnt[col_end_bbnt].apply(extract_endpoint)
        df_tk['_key'] = df_tk[col_obj_tk].apply(extract_endpoint)

        # Merge - Left Join on BBNT to keep list of Cables
        # "Dữ liệu hiện ra tab tuyến cáp được lấy từ cột tuyến cáp (cột B) trong file tuyen_cap"
        merged = pd.merge(df_bbnt, df_tk, on='_key', how='left', suffixes=('_BBNT', '_TK'))
        
        # Helper for resolving merged column names
        def _g(r, col, suffix):
            # Check original column name
            if col in r: return r[col]
            # Check suffixed column name
            if str(col)+suffix in r: return r[str(col)+suffix]
            return None

        results = []
        for _, row in merged.iterrows():
            tuyen_name = _g(row, col_tuyen_bbnt, '_BBNT')
            start_val_raw = _g(row, col_start_bbnt, '_BBNT')
            start_val = extract_endpoint(start_val_raw) if pd.notna(start_val_raw) and str(start_val_raw).strip() != '' else ""
            
            # Skip empty rows if any
            if pd.isna(tuyen_name) or str(tuyen_name).strip() == '':
                continue
                
            errors = []
            
            # Check Match
            obj_tk_val = _g(row, col_obj_tk, '_TK')
            if pd.isna(obj_tk_val):  # If merge failed
                 pass 

            # 1. Capacity (F vs C)
            cap_bbnt = self._safe_num(_g(row, col_cap_bbnt, '_BBNT'))
            cap_tk = self._safe_num(_g(row, col_cap_tk, '_TK'))
            if cap_bbnt != cap_tk:
                errors.append(f"Lệch Dlượng: TT={cap_bbnt} vs TK={cap_tk}")

            # 2. Type (G vs D) -> Map Treo=CT, Ngam=CC
            type_bbnt_raw = self._normalize_text(row.get(col_type_bbnt))
            type_tk_raw = self._normalize_text(row.get(col_type_tk))
            
            # Mapping BBNT -> Design code
            mapped_type = type_bbnt_raw
            if "treo" in type_bbnt_raw: mapped_type = "ct"
            elif "ngam" in type_bbnt_raw or "ngầm" in type_bbnt_raw: mapped_type = "cc"
            
            # Allow partial match? e.g. "cap treo" vs "ct"
            if mapped_type != type_tk_raw:
                 # Check if TK is empty?
                 if not type_tk_raw and type_bbnt_raw:
                     errors.append(f"Thiếu Loại cáp TK (TT={type_bbnt_raw})")
                 elif type_tk_raw:
                     errors.append(f"Lệch Loại: TT={type_bbnt_raw} vs TK={type_tk_raw}")

            # 3. Length (I vs E)
            len_bbnt = self._safe_num(row.get(col_len_bbnt))
            len_tk = self._safe_num(row.get(col_len_tk))
            
            if abs(len_bbnt - len_tk) > 1:
                errors.append(f"Lệch Chiều dài: TT={len_bbnt} vs TK={len_tk}")

            status = "✅ Khớp"
            if errors: 
                status = "❌ Sai lệch"
            
            # If no Design found at all for this key?
            if pd.isna(row.get(col_cap_tk)) and pd.isna(row.get(col_len_tk)):
                 status = "⚠️ Không tìm thấy TK"
                 errors = [f"Không tìm thấy đối tượng '{row['_key']}' trong file TK"]

            results.append({
                "Tuyến cáp": tuyen_name,
                "Điểm đầu": start_val,
                "Điểm cuối (Key)": row['_key'],
                "Dung lượng (TT/TK)": f"{int(cap_bbnt)} / {int(cap_tk)}",
                "Loại (TT/TK)": f"{type_bbnt_raw} / {type_tk_raw}",
                "Chiều dài (TT/TK)": f"{len_bbnt:g} / {len_tk:g}",
                "Trạng thái Lỗi": status,
                "Chi tiết": "; ".join(errors)
            })
            
        return pd.DataFrame(results)

    def check_vat_tu(self, df_bbnt_vattu: pd.DataFrame, df_bbnt_dt: pd.DataFrame, df_bbnt_tuyencap: pd.DataFrame) -> pd.DataFrame:
        """
        4. Hàm check_vat_tu (Updated):
        - Bỏ gom nhóm, xử lý từng dòng trong BBNT Vật tư.
        - SL Thiết kế = Tổng SL thực tế (Cột J) file doi_tuong + Tổng SL thực tế (Cột E) file tuyen_cap theo Mã Vật Tư.
        - SL Thực tế = SL thực tế (Cột H) file vat_tu.
        - Hiển thị Tình trạng hàng, bỏ Đơn vị tính.
        """
        # --- 1. Xác định các cột trong df_bbnt_vattu (File Vật Tư) ---
        col_code_vt = self._find_column(df_bbnt_vattu, ["Mã vật tư", "Mã VT"])
        if not col_code_vt and len(df_bbnt_vattu.columns) > 1: col_code_vt = df_bbnt_vattu.columns[1]
        
        col_name_vt = self._find_column(df_bbnt_vattu, ["Tên vật tư", "Tên VT", "Nội dung", "Diễn giải"])
        if not col_name_vt and len(df_bbnt_vattu.columns) > 2: col_name_vt = df_bbnt_vattu.columns[2]
        
        # SL thực tế trong vat_tu (User: Cột H -> index 7)
        col_sl_nt = self._find_column(df_bbnt_vattu, ["SL Thực tế", "SL Nghiệm thu", "Khối lượng NT", "Thực tế"])
        if not col_sl_nt and len(df_bbnt_vattu.columns) > 7: col_sl_nt = df_bbnt_vattu.columns[7]
        
        # Tình trạng hàng trong vat_tu
        col_tinhtrang = self._find_column(df_bbnt_vattu, ["Tình trạng hàng", "Tình trạng", "Chất lượng"])
        if not col_tinhtrang and len(df_bbnt_vattu.columns) > 8: col_tinhtrang = df_bbnt_vattu.columns[8] # Fallback bừa nếu ko có

        if not col_code_vt:
             return pd.DataFrame({"Lỗi": ["Không tìm thấy cột 'Mã vật tư' trong file BBNT Vật tư"]})

        # --- 2. Tính tổng SL Thiết Kế từ doi_tuong và tuyen_cap ---
        design_caps = {} # Sum mapping: Code -> Total SL Thiết kế

        # Từ file doi_tuong: Mã vật tư ở cột B (index 1), SL thực tế bù đắp ở cột J (index 9)
        if df_bbnt_dt is not None and not df_bbnt_dt.empty:
            current_code = ""
            for i in range(len(df_bbnt_dt)):
                try:
                    code_raw = str(df_bbnt_dt.iloc[i, 1]).strip().upper()
                    # Bỏ qua dòng tiêu đề, rác
                    if code_raw == 'MÃ VẬT TƯ' or code_raw == 'MÃ VT':
                        continue

                    if code_raw and code_raw != 'NAN' and code_raw != 'NONE':
                        current_code = code_raw

                    if current_code:
                        val = self._safe_num(df_bbnt_dt.iloc[i, 9])
                        design_caps[current_code] = design_caps.get(current_code, 0) + val
                except IndexError:
                    pass # Bỏ qua nếu dòng không đủ cột

        # Từ file tuyen_cap: Mã vật tư ở cột B (index 1), SL thực tế ở cột E (index 4)
        if df_bbnt_tuyencap is not None and not df_bbnt_tuyencap.empty:
            current_code = ""
            for i in range(len(df_bbnt_tuyencap)):
                try:
                    code_raw = str(df_bbnt_tuyencap.iloc[i, 1]).strip().upper()
                    if code_raw == 'MÃ VẬT TƯ' or code_raw == 'MÃ VT':
                        continue

                    if code_raw and code_raw != 'NAN' and code_raw != 'NONE':
                        current_code = code_raw

                    if current_code:
                        val = self._safe_num(df_bbnt_tuyencap.iloc[i, 4])
                        design_caps[current_code] = design_caps.get(current_code, 0) + val
                except IndexError:
                    pass

        # --- 3. Duyệt từng dòng trong BBNT Vật tư và Gom nhóm ---
        grouped_vt = {} # Code -> {name, sl_nt, tinh_trang_list}
        for _, row in df_bbnt_vattu.iterrows():
            code_raw = row.get(col_code_vt)
            if pd.isna(code_raw) or str(code_raw).strip() == '': continue
            
            code = str(code_raw).strip().upper()
            if code == 'NAN': continue
            
            name = str(row.get(col_name_vt, ''))
            tinh_trang = str(row.get(col_tinhtrang, '')) if col_tinhtrang else ''
            if pd.isna(row.get(col_tinhtrang)): tinh_trang = ''
            
            s_nt = self._safe_num(row.get(col_sl_nt, 0))
            
            if code not in grouped_vt:
                grouped_vt[code] = {
                    "name": name,
                    "sl_nt": s_nt,
                    "tinh_trang": [tinh_trang] if tinh_trang else []
                }
            else:
                grouped_vt[code]["sl_nt"] += s_nt
                if tinh_trang and tinh_trang not in grouped_vt[code]["tinh_trang"]:
                    grouped_vt[code]["tinh_trang"].append(tinh_trang)

        results = []
        for code, data in grouped_vt.items():
            name = data["name"]
            s_nt = data["sl_nt"]
            tinh_trang_str = ", ".join(data["tinh_trang"])
            s_tk = design_caps.get(code, 0.0)
            
            # Compare
            status = "✅ Khớp"
            details = []
            
            diff = s_tk - s_nt
            if abs(diff) > 0.001:
                status = "❌ Lệch SL"
                details.append(f"TK={s_tk:g} - NT={s_nt:g} => Lệch {diff:g}")

            results.append({
                "Mã vật tư": code,
                "Tên vật tư": name,
                "Tình trạng hàng": tinh_trang_str,
                "SL Thiết kế": round(s_tk, 1),
                "SL Nghiệm thu": round(s_nt, 1),
                "Trạng thái Lỗi": status,
                "Chi tiết": "; ".join(details)
            })
             
        if not results:
            return pd.DataFrame({"Lỗi": ["Không tìm thấy dữ liệu hợp lệ trong file BBNT Vật tư"]})
            
        return pd.DataFrame(results)

    def check_design_capacity(self, df_imp: pd.DataFrame, df_tk: pd.DataFrame) -> pd.DataFrame:
        """
        New Check: So sánh Form_import và thiet_ke (Dung lượng).
        - Imp: Tổng SP1.x, SP2.x (số port có dữ liệu).
        - Tk: Parse 'Loại bộ chia' (1:16 -> 16, 1:8 -> 8).
        """
        # --- 1. Calculate Import Usage ---
        # Find columns SP1.* and SP2.*
        sp1_cols = [c for c in df_imp.columns if 'sp1.' in str(c).lower()]
        sp2_cols = [c for c in df_imp.columns if 'sp2.' in str(c).lower()]
        
        # Calculate used ports
        def count_used(row, cols):
            cnt = 0
            for c in cols:
                val = str(row[c]).strip().lower()
                if val and val != 'nan' and val != '0' and val != 'none':
                    cnt += 1
            return cnt

        df_imp = df_imp.copy()
        df_imp['Used_SP1'] = df_imp.apply(lambda r: count_used(r, sp1_cols), axis=1)
        df_imp['Used_SP2'] = df_imp.apply(lambda r: count_used(r, sp2_cols), axis=1)

        key_imp = self._find_column(df_imp, ["Mã đối tượng", "Mã ĐT"])
        if not key_imp: return pd.DataFrame({"Lỗi": ["Không tìm thấy Mã ĐT trong FormImport"]})
        
        df_imp['_key'] = df_imp[key_imp].apply(self._normalize_text)
        
        # --- 2. Get Design Capacity ---
        key_tk = self._find_column(df_tk, ["Mã đối tượng", "Mã ĐT"])
        col_type = self._find_column(df_tk, ["Loại bộ chia", "Loại thiết bị", "Loại"])
        
        if not key_tk: return pd.DataFrame({"Lỗi": ["Không tìm thấy Mã ĐT trong Thiết kế"]})
        
        df_tk = df_tk.copy()
        df_tk['_key'] = df_tk[key_tk].apply(self._normalize_text)
        
        # Parse Capacity from Type
        def parse_cap(val):
            s = str(val).lower()
            if '1:16' in s: return 16
            if '1:8' in s: return 8
            if '1:4' in s: return 4
            if '1:2' in s: return 2
            if '1:32' in s: return 32
            if '1:64' in s: return 64
            return 0 # Unknown
            
        cap_val = 0
        if col_type:
            df_tk['_DesignCap'] = df_tk[col_type].apply(parse_cap)
        else:
            df_tk['_DesignCap'] = 0

        # --- 3. Merge & Compare ---
        merged = pd.merge(df_imp, df_tk, on='_key', how='inner', suffixes=('_IMP', '_TK'))
        
        results = []
        for _, row in merged.iterrows():
            ma_dt = row['_key']
            used_1 = row['Used_SP1']
            used_2 = row['Used_SP2']
            design_cap = row.get('_DesignCap', 0)
            
            # Logic user requested:
            # "thiet_ke thì dung lượng theo nguyên tắc 1:16; là bộ chia cấp 1 16 port, 1:16 là bộ chia cấp 2 16 port"
            # So Design Cap applies to BOTH Level 1 and Level 2? 
            # Or depends on device type?
            # Assuming the 'Loại bộ chia' column defines the max capacity for THAT device instance.
            # If used > design => Error.
            
            status = "✅ Khớp"
            msgs = []
            
            # Check SP1 (Assuming this device handles SP1?)
            # Or is SP1/SP2 distinct devices?
            # Usually FormImport has one row per 'Splitter' representing one Box.
            # If Box has splits 1:16, max is 16.
            # Total used = SP1_used + SP2_used?
            # Or SP1 columns are for Level 1 splitters and SP2 for Level 2?
            # If the row is ONE Object, it usually is ONE Splitter level.
            # But if columns SP1... and SP2... exist in SAME row, maybe it lists connections?
            # Let's check max(Used_SP1, Used_SP2) vs Design?
            # Or Sum?
            # User said: "dung lượng bộ chia cấp 1 là tổng... SP1.x, cấp 2 là tổng ... SP2.x".
            # And Design "1:16 là bộ chia cấp 1... 1:16 là bộ chia cấp 2".
            # This implies the Design Type tells us if it is Lv1 or Lv2.
            # If Design says "1:16", and used (sum) > 16 => Error.
            # I will compare MAX(Used_SP1, Used_SP2) against DesignCap for safety, assumption is one device per row.
            
            total_used = used_1 + used_2
            # Refine: if SP1 and SP2 both present, maybe it's 2 devices?
            # But merge is 1:1.
            # I'll compare Total Used ports vs Design Cap.
            
            if design_cap > 0:
                if total_used > design_cap:
                    status = "❌ Quá tải (Used > Design)"
                    msgs.append(f"Used={total_used} > Design={design_cap}")
                elif total_used == 0:
                    status = "⚠️ Không có thuê bao (Empty)"
            else:
                 msgs.append("Không xác định được loại bộ chia trong TK")
            
            if status != "✅ Khớp":
                 results.append({
                    "Mã ĐT": ma_dt,
                    "Used SP1": used_1,
                    "Used SP2": used_2,
                    "Design Cap": design_cap,
                    "Trạng thái Lỗi": status,
                    "Chi tiết": "; ".join(msgs)
                })
        
        if not results:
             return pd.DataFrame({"Kết quả": ["OK - Không thấy lỗi dung lượng"]})
             
        return pd.DataFrame(results)
